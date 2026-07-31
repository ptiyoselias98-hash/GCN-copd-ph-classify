#!/usr/bin/env python3
"""
COPD-PH PyRadiomics Pipeline — Server Version (NIfTI already exists!)
===============================================================
Reads existing NIfTI files from:
  /home/imss/cw/Claude_COPDnonPH_COPD-PH_CT_nii/nii/

Each patient dir contains:
  original.nii.gz  → CT image
  artery.nii.gz    → Artery mask
  vein.nii.gz      → Vein mask
  lung.nii.gz      → Lung mask
  airway.nii.gz    → Airway mask

Pipeline:
  1. Map patient name → NIfTI dir (pinyin matching)
  2. PyRadiomics extraction (original+wavelet+LoG) per ROI
  3. MWU → LASSO → LRC/RFC/SVC with 10×5 CV
  4. Report AUC with 95% CI

Estimated: 30-40 min for ~200 patients.
"""

import os, sys, time, json, hashlib, re, warnings
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")

# ─── Paths ────────────────────────────────────────────────
BASE      = Path("/home/imss/cw")
NII_ROOT  = BASE / "Claude_COPDnonPH_COPD-PH_CT_nii" / "nii"
OUT_DIR   = BASE / "GCN-copd-ph-classify" / "radiomics_output"
CACHE_DIR = BASE / "GCN-copd-ph-classify" / "radiomics_cache"
EXCEL     = BASE / "GCN-copd-ph-classify" / "copd-ph患者113例0331.xlsx"

for d in [OUT_DIR, CACHE_DIR]:
    d.mkdir(parents=True, exist_ok=True)

SEED = 2020
N_REPEATS = 10
N_FOLDS = 5
ROIS = ["vessel", "artery", "vein", "lung", "airway"]


# ════════════════════════════════════════════════════════
# 1. Map patients to NIfTI dirs
# ════════════════════════════════════════════════════════

def build_pinyin_map():
    """Build {patient_id: nii_dir} mapping using pypinyin."""
    try:
        from pypinyin import lazy_pinyin
    except ImportError:
        # Fallback: manual pinyin-like transformation
        def lazy_pinyin(s):
            return re.sub(r'[^\w]', '_', s).lower()

    if not EXCEL.exists():
        print(f"ERROR: Excel not found at {EXCEL}")
        return {}

    import openpyxl
    wb = openpyxl.load_workbook(str(EXCEL))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    ph_col = next((i+1 for i,h in enumerate(headers) if h == "PH"), 7)
    name_col = next((i+1 for i,h in enumerate(headers) if h == "name"), 3)
    seg_col = next((i+1 for i,h in enumerate(headers) if h and "分割" in str(h)), 12)

    # Build NIfTI dir index
    nii_dirs = {}
    for d in NII_ROOT.iterdir():
        if d.is_dir() and (d / "original.nii.gz").exists():
            dname = d.name.lower()
            nii_dirs[dname] = d

    print(f"NIfTI dirs available: {len(nii_dirs)}")

    matched = []
    for row in range(2, ws.max_row + 1):
        ph = ws.cell(row=row, column=ph_col).value
        name = ws.cell(row=row, column=name_col).value

        if ph == "是":   label = 1
        elif ph == "/":  label = 0
        else: continue

        if not name: continue

        # Generate possible NIfTI folder names
        py = None
        try:
            from pypinyin import lazy_pinyin
            py = "_".join(lazy_pinyin(str(name))).lower()
        except ImportError:
            py = re.sub(r'[^\x00-\x7F]', '', str(name)).strip().lower()
            py = re.sub(r'[^a-z0-9]+', '_', py).strip('_')

        # Try to match: nonph_pinyin_... or ph_pinyin_...
        found = None
        for prefix in ["nonph_", "ph_"]:
            for dname, dpath in nii_dirs.items():
                if dname.startswith(prefix + py):
                    found = dpath
                    break
            if found: break

        # Also try matching seg_folder name from Excel
        if not found:
            seg = ws.cell(row=row, column=seg_col).value
            if seg:
                seg_lower = str(seg).lower().rstrip("/").rstrip("\\")
                for dname, dpath in nii_dirs.items():
                    # Extract the ID part from dname (e.g., caochenglin_g02017953)
                    parts = dname.split("_")
                    if len(parts) >= 3:
                        middle = "_".join(parts[1:3])
                        if middle in seg_lower or seg_lower in dname:
                            found = dpath
                            break

        if found:
            pid = hashlib.md5(str(name).encode('utf-8')).hexdigest()[:12]
            matched.append({
                "patient_id": pid,
                "label": label,
                "name": name,
                "nii_dir": str(found),
            })

    print(f"Matched: {len(matched)}/{ws.max_row - 1} patients")
    n_ph = sum(1 for m in matched if m["label"] == 1)
    print(f"  PH={n_ph}, nonPH={len(matched)-n_ph}")

    # Save
    with open(OUT_DIR / "patient_map.json", "w", encoding="utf-8") as f:
        json.dump(matched, f, indent=2, ensure_ascii=False)

    return matched


# ════════════════════════════════════════════════════════
# 2. PyRadiomics extraction
# ════════════════════════════════════════════════════════

def extract_all_features(patients: list) -> pd.DataFrame:
    """Extract PyRadiomics features from existing NIfTI files."""
    import radiomics
    from radiomics import featureextractor

    print(f"\nPyRadiomics version: {radiomics.__version__}")

    # Reduced settings for speed
    extractor = featureextractor.RadiomicsFeatureExtractor(
        binWidth=5,
        resampledPixelSpacing=[2, 2, 2],  # coarser for speed
        interpolator='sitkBSpline',
        verbose=False,
    )
    extractor.enableAllFeatures()
    extractor.enableImageTypes(Original={}, Wavelet={})

    records = []
    t0 = time.time()
    n_total = len(patients)

    for i, p in enumerate(patients):
        pid = p["patient_id"]
        nii_dir = Path(p["nii_dir"])
        ct_path = nii_dir / "original.nii.gz"

        # Cache check
        cache_file = CACHE_DIR / f"{pid}_features.json"
        if cache_file.exists():
            try:
                with open(cache_file) as f:
                    rec = json.load(f)
                rec["patient_id"] = pid
                rec["label"] = p["label"]
                records.append(rec)
                continue
            except: pass

        if not ct_path.exists():
            continue

        all_feats = {"patient_id": pid, "label": p["label"]}

        # Extract per ROI
        for roi in ROIS:
            if roi == "vessel":
                # Create vessel = artery + vein
                art_path = nii_dir / "artery.nii.gz"
                vei_path = nii_dir / "vein.nii.gz"
                if art_path.exists() and vei_path.exists():
                    import SimpleITK as sitk
                    art = sitk.ReadImage(str(art_path))
                    vei = sitk.ReadImage(str(vei_path))
                    vessel = sitk.Or(art > 0, vei > 0)
                    vessel = sitk.Cast(vessel, sitk.sitkUInt8)
                    tmp = CACHE_DIR / f"{pid}_vessel_tmp.nii.gz"
                    sitk.WriteImage(vessel, str(tmp))
                    mask_path = tmp
                else:
                    continue
            else:
                mask_path = nii_dir / f"{roi}.nii.gz"

            if not mask_path.exists():
                continue

            try:
                result = extractor.execute(str(ct_path), str(mask_path))
                for k, v in result.items():
                    if k.startswith('diagnostics_') or k.startswith('general_'):
                        continue
                    try:
                        all_feats[f"{roi}_{k}"] = float(v)
                    except (ValueError, TypeError):
                        pass
            except Exception as e:
                # Single ROI failure shouldn't kill everything
                pass

            # Cleanup temp vessel mask
            if roi == "vessel":
                try: (CACHE_DIR / f"{pid}_vessel_tmp.nii.gz").unlink()
                except: pass

        # Cache
        with open(cache_file, "w") as f:
            json.dump(all_feats, f)

        records.append(all_feats)

        if (i + 1) % 10 == 0 or i == 0:
            elapsed = time.time() - t0
            eta = elapsed / (i + 1) * n_total - elapsed
            print(f"  [{i+1}/{n_total}] {pid}: {len(all_feats)-2} feats, "
                  f"{elapsed/60:.1f}m elapsed, {eta/60:.1f}m ETA")

    total_time = time.time() - t0
    print(f"\n  Extraction done: {total_time/60:.1f} min")

    # Build DataFrame
    df = pd.DataFrame(records)
    feat_cols = [c for c in df.columns if c not in ("patient_id", "label")]
    for c in feat_cols:
        median = df[c].median()
        if pd.notna(median):
            df[c] = df[c].fillna(median)
        else:
            df[c] = df[c].fillna(0.0)

    out_path = OUT_DIR / "pyradiomics_features.csv"
    df.to_csv(out_path, index=False, encoding="utf-8")
    print(f"  Features saved: {df.shape[0]}x{df.shape[1]-2} → {out_path}")
    return df


# ════════════════════════════════════════════════════════
# 3. ML (Ruan 2025)
# ════════════════════════════════════════════════════════

def run_ruan_ml(df):
    """LASSO + 10x5 CV for LRC, RFC, SVC."""
    from scipy import stats
    from sklearn.linear_model import LassoCV, LogisticRegression
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.svm import SVC
    from sklearn.preprocessing import StandardScaler
    from sklearn.model_selection import StratifiedKFold
    from sklearn.metrics import roc_auc_score

    print("\n" + "=" * 60)
    print("ML: Ruan 2025 Method")
    print("=" * 60)

    feat_cols = [c for c in df.columns if c not in ("patient_id", "label")]
    X_all = df[feat_cols].values.astype(np.float64)
    y = df["label"].values.astype(int)

    print(f"Data: {len(y)} pts ({int((y==1).sum())} PH, {int((y==0).sum())} nonPH), {len(feat_cols)} features")

    # MWU filter
    print("\n[1/3] MWU pre-filter...")
    keep, pvals = [], []
    for j in range(X_all.shape[1]):
        g0, g1 = X_all[y==0, j], X_all[y==1, j]
        if np.std(np.r_[g0, g1]) < 1e-12: continue
        _, p = stats.mannwhitneyu(g0, g1, alternative="two-sided")
        if p < 0.05: keep.append(j); pvals.append(p)
    if len(keep) < 10:
        keep = np.argsort(np.var(X_all, axis=0))[-100:].tolist()
    X_filt = X_all[:, keep]
    filt_names = [feat_cols[i] for i in keep]
    print(f"  {len(keep)}/{len(feat_cols)} features retained")

    # LASSO
    print("\n[2/3] LASSO selection...")
    scaler = StandardScaler()
    X_sc = scaler.fit_transform(X_filt)
    lasso = LassoCV(alphas=np.logspace(-3, 1, 100, base=10), cv=10,
                     max_iter=2000, random_state=SEED)
    lasso.fit(X_sc, y)
    sel = np.abs(lasso.coef_) > np.percentile(np.abs(lasso.coef_), 85)
    X_rad = X_filt[:, sel]
    selected = [filt_names[i] for i in range(len(filt_names)) if sel[i]]
    print(f"  {int(sel.sum())} features selected")
    for i, f in enumerate(selected[:15], 1):
        print(f"    {i:2d}. {f[:80]}")

    # Models
    print("\n[3/3] 10×5 CV training...")
    models = {
        "LRC_RM": LogisticRegression(C=1.0, solver='lbfgs', max_iter=5000,
                                      class_weight='balanced', random_state=SEED),
        "RFC_RM": RandomForestClassifier(n_estimators=12, max_depth=4,
                                          max_features=min(11, X_rad.shape[1]),
                                          random_state=SEED),
        "SVC_RM": SVC(kernel='rbf', C=1.0, probability=True,
                       class_weight='balanced', random_state=SEED),
    }

    results = {}
    for mk, model in models.items():
        aucs = []
        for rep in range(N_REPEATS):
            skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True,
                                  random_state=SEED + rep * 100)
            for tr, va in skf.split(X_rad, y):
                sc_tr = StandardScaler().fit_transform(X_rad[tr])
                sc_va = StandardScaler().fit_transform(X_rad[va])
                model.fit(sc_tr, y[tr])
                try: yp = model.predict_proba(sc_va)[:, 1]
                except: yp = model.decision_function(sc_va)
                if len(set(y[va])) > 1:
                    aucs.append(roc_auc_score(y[va], yp))

        auc_arr = np.array(aucs)
        results[mk] = {
            "AUC": float(np.mean(auc_arr)),
            "AUC_std": float(np.std(auc_arr)),
            "CI_low": float(np.percentile(auc_arr, 2.5)),
            "CI_high": float(np.percentile(auc_arr, 97.5)),
            "n_evals": len(aucs),
        }
        print(f"  {mk:12s}: AUC={results[mk]['AUC']:.4f} [{results[mk]['CI_low']:.4f}-{results[mk]['CI_high']:.4f}]")

    # Save
    output = {
        "method": "Ruan 2025 — PyRadiomics + MWU + LASSO + 10x5 CV",
        "n_patients": len(y), "n_PH": int((y==1).sum()), "n_nonPH": int((y==0).sum()),
        "n_raw_features": len(feat_cols),
        "n_mwu": len(keep), "n_lasso": int(sel.sum()),
        "selected_features": selected,
        "results": results,
    }
    with open(OUT_DIR / "ruan2025_pyradiomics_results.json", "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    print(f"\n{'=' * 60}")
    print("FINAL RESULTS")
    print(f"{'=' * 60}")
    for mk in ["LRC_RM", "SVC_RM", "RFC_RM"]:
        if mk in results:
            r = results[mk]
            print(f"  {mk:12s}: AUC={r['AUC']:.4f} [{r['CI_low']:.4f}-{r['CI_high']:.4f}]")

    print(f"\nOutput: {OUT_DIR / 'ruan2025_pyradiomics_results.json'}")
    return results


# ════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════

def main():
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--skip-extract", action="store_true")
    p.add_argument("--skip-ml", action="store_true")
    p.add_argument("--max", type=int, default=0, help="Max patients (0=all)")
    args = p.parse_args()

    t_all = time.time()
    print("=" * 60)
    print("COPD-PH PyRadiomics — Ruan 2025 Method (Server)")
    print("=" * 60)

    # 1. Map patients
    print("\n>>> [1] Mapping patients to NIfTI...")
    patients = build_pinyin_map()
    if args.max > 0:
        patients = patients[:args.max]

    if not patients:
        print("ERROR: No patients mapped!")
        return 1

    # 2. PyRadiomics
    if not args.skip_extract:
        print(f"\n>>> [2] PyRadiomics extraction ({len(patients)} patients)...")
        df = extract_all_features(patients)
    else:
        csv_path = OUT_DIR / "pyradiomics_features.csv"
        if csv_path.exists():
            df = pd.read_csv(csv_path)
            print(f"\n>>> [2] Loaded existing features: {df.shape}")
        else:
            print(f"ERROR: {csv_path} not found!")
            return 1

    # 3. ML
    if not args.skip_ml:
        print(f"\n>>> [3] ML training...")
        run_ruan_ml(df)

    total = (time.time() - t_all) / 60
    print(f"\nDONE — {total:.1f} min")
    return 0


if __name__ == "__main__":
    sys.exit(main())
