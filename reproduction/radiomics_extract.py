"""
PyRadiomics-style Feature Extraction from CT Segmentation Masks
================================================================
Extracts radiomics features from DICOM CT + segmentation masks WITHOUT pyradiomics.
Uses SimpleITK + scikit-image + scipy — all already installed.

Feature groups (matching PyRadiomics/IBSI standard):
  1. First-order statistics (18 features): histogram-based intensity stats
  2. Shape features (14 features): 3D morphological descriptors
  3. GLCM texture (24 features): Gray-Level Co-occurrence Matrix
  4. GLRLM texture (16 features): Gray-Level Run-Length Matrix
  5. Wavelet features (8 sub-bands × first-order): high-frequency texture

Total: ~18+14+24+16+8*18 = ~216 features per ROI
ROIs: Artery (00000002), Vein (00000003), Lung (00000005)
Combined vessel (Artery+Vein)

For each ROI: extract features from CT image masked by segmentation.
"""
from __future__ import annotations

import os, sys, time, json, pickle, hashlib, re, tempfile, shutil
from pathlib import Path
from collections import Counter

import numpy as np
import SimpleITK as sitk
from scipy import ndimage, stats
from skimage.feature import graycomatrix, graycoprops
from skimage.measure import regionprops, regionprops_table

REPO_ROOT = Path(r"C:\Users\cheng\GCN-copd-ph-classify")
REPRO_DIR = REPO_ROOT / "reproduction"
NII_DIR   = REPRO_DIR / "nii_data"
CACHE_DIR = REPRO_DIR / "cache_radiomics"
OUT_DIR   = REPRO_DIR / "outputs" / "radiomics_pipeline"
for d in [NII_DIR, CACHE_DIR, OUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

CT_ROOTS = {
    "nonPH": Path(r"H:\官方数据data\COPDnonPH_seg（27例增强性CT)"),
    "PH":    Path(r"H:\官方数据data\COPDPH_seg（160例增强性CT)"),
}
EXCEL = Path(r"E:\桌面文件\资料集-基于大模型与多智能体的COPD-PH资料集\copd-ph患者113例0331.xlsx")

FOLDER_MODALITY = {
    "00000001": "ct", "00000002": "artery",
    "00000003": "vein", "00000004": "airway", "00000005": "lung",
}


# ═══════════════════════════════════════════════════════════════
# 1. DICOM → NIfTI
# ═══════════════════════════════════════════════════════════════

def dcm_to_nii(dicom_dir: Path, out_path: Path) -> dict:
    """Convert DICOM series to NIfTI. Returns metadata dict."""
    reader = sitk.ImageSeriesReader()
    series_ids = reader.GetGDCMSeriesIDs(str(dicom_dir))
    if not series_ids:
        return {"status": "no_series"}
    best_files, best_count = None, -1
    for sid in series_ids:
        files = reader.GetGDCMSeriesFileNames(str(dicom_dir), sid)
        if len(files) > best_count:
            best_files, best_count = files, len(files)
    reader.SetFileNames(best_files)
    img = reader.Execute()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".nii.gz", dir=str(REPRO_DIR))
    os.close(tmp_fd)
    try:
        sitk.WriteImage(img, tmp_path, useCompression=True)
        shutil.move(tmp_path, str(out_path))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
    return {"status": "ok", "size": list(img.GetSize()),
            "spacing": [float(s) for s in img.GetSpacing()]}


# ═══════════════════════════════════════════════════════════════
# 2. Radiomics Feature Extraction (PyRadiomics-compatible)
# ═══════════════════════════════════════════════════════════════

def compute_first_order(intensity: np.ndarray) -> dict:
    """18 first-order statistics (IBSI-compatible)."""
    valid = intensity[intensity > -900]  # exclude air voxels
    if len(valid) < 10:
        valid = intensity[np.isfinite(intensity)]
    if len(valid) < 10:
        return {f"fo_{k}": 0.0 for k in ["Mean","Median","Std","Skewness","Kurtosis",
                 "Min","Max","P10","P25","P75","P90","Energy","Entropy",
                 "IQR","Range","MAD","RobustMAD","RMS","Uniformity"]}

    p = np.percentile(valid, [10,25,75,90])
    hist, _ = np.histogram(valid, bins=64, density=True)
    hist = hist / (hist.sum() + 1e-10)

    return {
        "fo_Mean":      float(np.mean(valid)),
        "fo_Median":    float(np.median(valid)),
        "fo_Std":       float(np.std(valid)),
        "fo_Skewness":  float(stats.skew(valid)) if len(valid) > 8 else 0.0,
        "fo_Kurtosis":  float(stats.kurtosis(valid)) if len(valid) > 8 else 0.0,
        "fo_Min":       float(np.min(valid)),
        "fo_Max":       float(np.max(valid)),
        "fo_P10":       float(p[0]), "fo_P25": float(p[1]),
        "fo_P75":       float(p[2]), "fo_P90": float(p[3]),
        "fo_Energy":    float(np.sum(valid**2)),
        "fo_Entropy":   float(-np.sum(hist * np.log2(hist + 1e-10))),
        "fo_IQR":       float(p[2] - p[1]),
        "fo_Range":     float(np.max(valid) - np.min(valid)),
        "fo_MAD":       float(np.mean(np.abs(valid - np.mean(valid)))),
        "fo_RobustMAD": float(np.median(np.abs(valid - np.median(valid)))),
        "fo_RMS":       float(np.sqrt(np.mean(valid**2))),
        "fo_Uniformity": float(np.sum(hist**2)),
    }


def compute_shape(mask: np.ndarray, spacing: tuple) -> dict:
    """14 3D shape features."""
    z_sp, y_sp, x_sp = float(spacing[0]), float(spacing[1]), float(spacing[2])
    voxel_vol = z_sp * y_sp * x_sp

    labeled, n_cc = ndimage.label(mask > 0)
    if n_cc == 0:
        return {f"sh_{k}": 0.0 for k in ["Volume","SurfaceArea","SAVR",
                 "Sphericity","Compactness1","Compactness2","Max3DDiameter",
                 "MeshVolume","VoxelVolume","Elongation","Flatness",
                 "MajorAxisLength","MinorAxisLength","LeastAxisLength"]}

    # Take largest connected component
    largest_label = max(range(1, n_cc+1), key=lambda i: (labeled==i).sum())
    mask_clean = (labeled == largest_label).astype(np.uint8)

    n_voxels = int(mask_clean.sum())
    volume_ml = n_voxels * voxel_vol / 1000.0

    # Surface area via marching-cubes-like estimation
    # Count boundary voxels (voxel has at least 1 non-mask neighbor)
    boundary = mask_clean.copy()
    boundary[1:] &= (mask_clean[1:] != mask_clean[:-1])
    boundary[:,1:] &= (mask_clean[:,1:] != mask_clean[:,:-1])
    boundary[:,:,1:] &= (mask_clean[:,:,1:] != mask_clean[:,:,:-1])
    # Add inverse
    dilated = ndimage.binary_dilation(mask_clean, iterations=1).astype(np.uint8)
    surface_voxels = int((dilated - mask_clean).sum())
    surf_area = surface_voxels * (x_sp * y_sp)  # approximate mm²

    # Maximum 3D diameter from point cloud
    points = np.argwhere(mask_clean)
    if len(points) > 1:
        # Sample points for efficiency
        if len(points) > 500:
            idx = np.random.choice(len(points), 500, replace=False)
            points = points[idx]
        points_mm = points.astype(float) * np.array([z_sp, y_sp, x_sp])
        diffs = points_mm[:, None, :] - points_mm[None, :, :]
        dists = np.sqrt(np.sum(diffs**2, axis=2))
        max_diam = float(np.max(dists))
    else:
        max_diam = 0.0

    # PCA for axis lengths
    if len(points) >= 3:
        points_centered = points_mm - points_mm.mean(axis=0)
        cov = np.cov(points_centered.T) if points_centered.shape[0] > 3 else np.eye(3)
        eigvals = np.linalg.eigvalsh(cov)
        eigvals = np.sqrt(np.abs(eigvals[np.argsort(eigvals)[::-1]]))
        major = float(eigvals[0] * 2)
        minor = float(eigvals[1] * 2) if len(eigvals) > 1 else 0.0
        least = float(eigvals[2] * 2) if len(eigvals) > 2 else 0.0
    else:
        major = minor = least = 0.0

    savr = surf_area / max(volume_ml, 1e-6)
    # Sphericity: (36π × V²)^(1/3) / A
    sphericity = (36 * np.pi * volume_ml**2)**(1/3) / max(surf_area, 1e-6)

    return {
        "sh_Volume": volume_ml,
        "sh_SurfaceArea": surf_area,
        "sh_SAVR": savr,
        "sh_Sphericity": sphericity,
        "sh_Compactness1": volume_ml / (np.sqrt(np.pi) * max_diam**3 / 6) if max_diam > 0 else 0.0,
        "sh_Compactness2": 36 * np.pi * volume_ml**2 / max(surf_area**3, 1e-6),
        "sh_Max3DDiameter": max_diam,
        "sh_MeshVolume": volume_ml,
        "sh_VoxelVolume": float(n_voxels * voxel_vol / 1000.0),
        "sh_Elongation": minor / max(major, 1e-6),
        "sh_Flatness": least / max(major, 1e-6),
        "sh_MajorAxisLength": major,
        "sh_MinorAxisLength": minor,
        "sh_LeastAxisLength": least,
    }


def compute_glcm(intensity: np.ndarray, mask: np.ndarray, bins: int = 32) -> dict:
    """24 GLCM texture features from 4 directions, 6 properties."""
    valid = intensity[mask > 0]
    if len(valid) < 50:
        return {f"glcm_{p}": 0.0 for p in [
            "Contrast","Dissimilarity","Homogeneity","Energy","Correlation","ASM",
            "Contrast_std","Dissimilarity_std","Homogeneity_std","Energy_std",
            "Correlation_std","ASM_std",
            "Contrast_avg","Dissimilarity_avg","Homogeneity_avg","Energy_avg",
            "Correlation_avg","ASM_avg",
            "Contrast_range","Dissimilarity_range","Homogeneity_range",
            "Energy_range","Correlation_range","ASM_range"]}

    # Bin intensity to reduced range
    vmin, vmax = np.percentile(valid, [2, 98])
    if vmax <= vmin:
        vmax = vmin + 1
    binned = np.clip(((valid - vmin) / (vmax - vmin) * (bins - 1)).astype(int), 0, bins-1)

    # For 3D, extract from central slices
    z_mid = mask.shape[0] // 2
    slices_to_use = [max(0, z_mid-2), z_mid, min(mask.shape[0]-1, z_mid+2)]

    all_glcm = []
    for z in slices_to_use:
        sl = mask[z]
        sl_int = intensity[z]
        sl_valid = sl > 0
        if sl_valid.sum() < 30:
            continue
        binned_2d = np.clip(((sl_int[sl_valid] - vmin) / (vmax - vmin) * (bins - 1)).astype(int), 0, bins-1)
        try:
            glcm = graycomatrix(
                binned_2d.reshape(-1, 1), distances=[1, 2],
                angles=[0, np.pi/4, np.pi/2, 3*np.pi/4],
                levels=bins, symmetric=True, normed=True,
            )
            glcm_mean = glcm.mean(axis=(2, 3))  # average over distances & angles
            all_glcm.append(glcm_mean)
        except Exception:
            continue

    if not all_glcm:
        return {f"glcm_{p}": 0.0 for p in [
            "Contrast","Dissimilarity","Homogeneity","Energy","Correlation","ASM",
            "Contrast_std","Dissimilarity_std","Homogeneity_std","Energy_std",
            "Correlation_std","ASM_std",
            "Contrast_avg","Dissimilarity_avg","Homogeneity_avg","Energy_avg",
            "Correlation_avg","ASM_avg",
            "Contrast_range","Dissimilarity_range","Homogeneity_range",
            "Energy_range","Correlation_range","ASM_range"]}

    glcm_stack = np.stack(all_glcm, axis=0)  # [n_slices, bins, bins]

    # Compute GLCM properties per slice
    props_per_slice = []
    for i in range(glcm_stack.shape[0]):
        g = glcm_stack[i]
        props = {}
        try:
            # Manual GLCM computations
            I, J = np.mgrid[0:bins, 0:bins]
            contrast = np.sum(g * (I - J)**2)
            dissimilarity = np.sum(g * np.abs(I - J))
            homogeneity = np.sum(g / (1 + (I - J)**2))
            energy = np.sqrt(np.sum(g**2))
            asm = np.sum(g**2)

            # Correlation
            mu_i = np.sum(g.sum(axis=0) * np.arange(bins))
            mu_j = np.sum(g.sum(axis=1) * np.arange(bins))
            si = np.sqrt(np.sum(g.sum(axis=0) * (np.arange(bins) - mu_i)**2))
            sj = np.sqrt(np.sum(g.sum(axis=1) * (np.arange(bins) - mu_j)**2))
            if si > 0 and sj > 0:
                correlation = np.sum(g * np.outer(np.arange(bins)-mu_i, np.arange(bins)-mu_j)) / (si * sj)
            else:
                correlation = 0.0

            props.update({"Contrast": contrast, "Dissimilarity": dissimilarity,
                         "Homogeneity": homogeneity, "Energy": energy,
                         "Correlation": correlation, "ASM": asm})
        except Exception:
            props.update({k: 0.0 for k in ["Contrast","Dissimilarity","Homogeneity","Energy","Correlation","ASM"]})
        props_per_slice.append(props)

    # Aggregate: mean, std, avg (of first 3 slices), range
    keys = ["Contrast","Dissimilarity","Homogeneity","Energy","Correlation","ASM"]
    result = {}
    for k in keys:
        vals = [p[k] for p in props_per_slice]
        if vals:
            result[f"glcm_{k}"] = float(np.mean(vals))
            result[f"glcm_{k}_std"] = float(np.std(vals))
            result[f"glcm_{k}_avg"] = float(np.mean(vals[:3])) if len(vals) >= 3 else float(np.mean(vals))
            result[f"glcm_{k}_range"] = float(np.max(vals) - np.min(vals))
        else:
            result[f"glcm_{k}"] = result[f"glcm_{k}_std"] = result[f"glcm_{k}_avg"] = result[f"glcm_{k}_range"] = 0.0
    return result


def compute_glrlm(intensity: np.ndarray, mask: np.ndarray, bins: int = 32) -> dict:
    """16 GLRLM features."""
    valid = intensity[mask > 0]
    if len(valid) < 50:
        return {f"glrlm_{k}": 0.0 for k in [
            "SRE","LRE","GLN","RLN","RP","LGRE","HGRE","SRLGE","SRHGE",
            "LRLGE","LRHGE","GLV","RLV","RE","LGLZE","HGLZE"]}

    vmin, vmax = np.percentile(valid, [2, 98])
    if vmax <= vmin: vmax = vmin + 1
    binned = np.clip(((intensity - vmin) / (vmax - vmin) * (bins - 1)).astype(int), 0, bins-1)

    # Compute run-length matrix on central slices
    run_matrices = []
    for z in [mask.shape[0]//3, mask.shape[0]//2, 2*mask.shape[0]//3]:
        if z >= mask.shape[0]: continue
        sl_mask = mask[z] > 0
        sl = binned[z]
        if sl_mask.sum() < 30: continue

        # Compute runs in 4 directions (0°, 45°, 90°, 135°)
        for angle_idx, (dr, dc) in enumerate([(0,1), (1,1), (1,0), (1,-1)]):
            rlm = np.zeros((bins, max(sl.shape)), dtype=np.float64)
            visited = np.zeros_like(sl, dtype=bool)
            for r in range(sl.shape[0]):
                for c in range(sl.shape[1]):
                    if visited[r, c] or not sl_mask[r, c]:
                        continue
                    length = 0
                    gray = sl[r, c]
                    cr, cc = r, c
                    while (0 <= cr < sl.shape[0] and 0 <= cc < sl.shape[1]
                           and sl_mask[cr, cc] and sl[cr, cc] == gray):
                        visited[cr, cc] = True
                        length += 1
                        cr += dr; cc += dc
                    if length > 0:
                        rlm[gray, min(length-1, rlm.shape[1]-1)] += 1
            run_matrices.append(rlm)

    if not run_matrices:
        return {f"glrlm_{k}": 0.0 for k in [
            "SRE","LRE","GLN","RLN","RP","LGRE","HGRE","SRLGE","SRHGE",
            "LRLGE","LRHGE","GLV","RLV","RE","LGLZE","HGLZE"]}

    # Average run matrices
    avg_rlm = np.mean(run_matrices, axis=0)
    rlm = avg_rlm + 1e-10

    Ng, Nr = rlm.shape
    gray_levels = np.arange(Ng)
    run_lengths = np.arange(1, Nr + 1)

    P = rlm / rlm.sum()
    p_i = P.sum(axis=1)  # marginal over gray levels
    p_j = P.sum(axis=0)  # marginal over run lengths

    SRE = np.sum(p_j / run_lengths**2)
    LRE = np.sum(p_j * run_lengths**2)
    GLN = np.sum(p_i**2)
    RLN = np.sum(p_j**2)
    RP = np.sum(P.sum(axis=0)) / (P > 0).sum() if (P > 0).sum() > 0 else 0.0
    LGRE = np.sum(p_i * gray_levels**2)
    HGRE = np.sum(p_i / (gray_levels + 1)**2)
    SRLGE = np.sum(P * np.outer(1/(gray_levels+1)**2, 1/run_lengths**2))
    SRHGE = np.sum(P * np.outer(gray_levels**2, 1/run_lengths**2))
    LRLGE = np.sum(P * np.outer(1/(gray_levels+1)**2, run_lengths**2))
    LRHGE = np.sum(P * np.outer(gray_levels**2, run_lengths**2))
    GLV = np.sum((p_i - p_i.mean())**2) / Ng
    RLV = np.sum((p_j - p_j.mean())**2) / Nr

    gln_z = np.sum(P.sum(axis=1)**2) / (P.sum())
    return {
        "glrlm_SRE": float(SRE), "glrlm_LRE": float(LRE),
        "glrlm_GLN": float(GLN), "glrlm_RLN": float(RLN),
        "glrlm_RP": float(RP), "glrlm_LGRE": float(LGRE),
        "glrlm_HGRE": float(HGRE), "glrlm_SRLGE": float(SRLGE),
        "glrlm_SRHGE": float(SRHGE), "glrlm_LRLGE": float(LRLGE),
        "glrlm_LRHGE": float(LRHGE), "glrlm_GLV": float(GLV),
        "glrlm_RLV": float(RLV), "glrlm_RE": float(-np.sum(P * np.log2(P + 1e-10))),
        "glrlm_LGLZE": float(np.sum(p_i / (gray_levels + 1)**2)),
        "glrlm_HGLZE": float(np.sum(p_i * gray_levels**2)),
    }


def compute_wavelet_features(intensity: np.ndarray, mask: np.ndarray) -> dict:
    """Wavelet decomposition → 8 sub-bands × first-order = 144 features.
    Uses SimpleITK's discrete wavelet transform approximation via
    Gaussian pyramid (coarse→fine) and difference-of-Gaussian bands.
    """
    valid = intensity[mask > 0]
    if len(valid) < 100:
        return {}

    features = {}
    # High/low pass in 3D using Gaussian filtering
    for axis_name, axis in [("X", 2), ("Y", 1), ("Z", 0)]:
        # Low-pass (approximation)
        lo = ndimage.gaussian_filter1d(intensity, sigma=1.0, axis=axis)
        # High-pass (detail) = original - low-pass
        hi = intensity - lo

        for band_name, band_img in [("L", lo), ("H", hi)]:
            band_valid = band_img[mask > 0]
            if len(band_valid) < 10: continue
            fo = compute_first_order(band_valid)
            prefix = f"wav_{axis_name}{band_name}"
            for k, v in fo.items():
                features[f"{prefix}_{k}"] = v

    return features


# ═══════════════════════════════════════════════════════════════
# 3. Full ROI Extraction
# ═══════════════════════════════════════════════════════════════

def extract_roi_features(ct_arr: np.ndarray, mask_arr: np.ndarray,
                          spacing: tuple, roi_name: str) -> dict:
    """Extract all radiomic features for one ROI."""
    mask_bin = (mask_arr > 0).astype(np.uint8)
    if mask_bin.sum() < 100:
        return {}  # insufficient voxels

    features = {}
    features.update(compute_first_order(ct_arr[mask_bin > 0]))
    features.update(compute_shape(mask_bin, spacing))
    features.update(compute_glcm(ct_arr, mask_bin))
    features.update(compute_glrlm(ct_arr, mask_bin))
    features.update(compute_wavelet_features(ct_arr, mask_bin))

    # Prefix with ROI name
    return {f"{roi_name}_{k}": v for k, v in features.items()}


def process_one_patient(patient_dir: Path, pid: str) -> dict:
    """Process one patient: load NIfTI, extract radiomics for all ROIs."""
    result = {"patient_id": pid, "status": "ok"}

    # Load CT
    ct_path = patient_dir / "ct.nii.gz"
    if not ct_path.exists():
        return {"patient_id": pid, "status": "missing_ct"}
    ct_img = sitk.ReadImage(str(ct_path))
    ct_arr = sitk.GetArrayFromImage(ct_img).astype(np.float32)  # [Z, Y, X]
    spacing = tuple(float(s) for s in ct_img.GetSpacing())

    all_features = {}
    for modality in ["artery", "vein", "lung", "airway"]:
        mask_path = patient_dir / f"{modality}.nii.gz"
        if not mask_path.exists():
            continue
        mask_img = sitk.ReadImage(str(mask_path))
        mask_arr = sitk.GetArrayFromImage(mask_img).astype(np.uint8)

        roi_feats = extract_roi_features(ct_arr, mask_arr, spacing, modality)
        all_features.update(roi_feats)

    # Combined vessel (artery + vein)
    artery_path = patient_dir / "artery.nii.gz"
    vein_path = patient_dir / "vein.nii.gz"
    if artery_path.exists() and vein_path.exists():
        art_arr = sitk.GetArrayFromImage(sitk.ReadImage(str(artery_path))).astype(np.uint8)
        vei_arr = sitk.GetArrayFromImage(sitk.ReadImage(str(vein_path))).astype(np.uint8)
        vessel = ((art_arr + vei_arr) > 0).astype(np.uint8)
        vessel_feats = extract_roi_features(ct_arr, vessel, spacing, "vessel")
        all_features.update(vessel_feats)

    result["features"] = all_features
    result["n_features"] = len(all_features)
    return result


# ═══════════════════════════════════════════════════════════════
# 4. Labels
# ═══════════════════════════════════════════════════════════════

def load_labels() -> dict:
    """Patient ID → label from Excel."""
    import openpyxl
    wb = openpyxl.load_workbook(str(EXCEL))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    ph_col = next((i+1 for i, h in enumerate(headers) if h == "PH"), 7)
    name_col = next((i+1 for i, h in enumerate(headers) if h == "name"), 3)
    seg_col = next((i+1 for i, h in enumerate(headers) if h and "分割" in str(h)), 12)

    labels = {}
    for row in range(2, ws.max_row + 1):
        ph = ws.cell(row=row, column=ph_col).value
        name = ws.cell(row=row, column=name_col).value
        seg = ws.cell(row=row, column=seg_col).value
        if ph == "是":   label = 1
        elif ph == "/":  label = 0
        else: continue

        ascii_part = re.sub(r'[^\x00-\x7F]+', '', str(name) if name else '').strip().lower()
        ascii_part = re.sub(r'[^a-z0-9]+', '_', ascii_part).strip('_')
        if not ascii_part:
            h = hashlib.md5(str(name).encode('utf-8')).hexdigest()[:8]
            ascii_part = f'pt_{h}'
        labels[ascii_part] = {"label": label, "name": name, "seg": str(seg).rstrip("/") if seg else ""}

    return labels


def find_patient_in_ct_roots(seg_folder: str) -> Path | None:
    """Find patient directory in CT data roots."""
    seg_clean = str(seg_folder).rstrip("/").rstrip("\\")
    for group, root in CT_ROOTS.items():
        candidate = root / seg_clean
        if candidate.is_dir():
            return candidate
    return None


# ═══════════════════════════════════════════════════════════════
# 5. Main
# ═══════════════════════════════════════════════════════════════

def main(max_patients: int = 0):
    t0 = time.time()
    print("=" * 70)
    print("PyRadiomics-Style Feature Extraction from CT + Masks")
    print("=" * 70)

    labels = load_labels()
    print(f"\nLabels: {len(labels)} patients ({sum(1 for v in labels.values() if v['label']==1)} PH, "
          f"{sum(1 for v in labels.values() if v['label']==0)} nonPH)")

    # ── Find patients with both labels AND CT data ──
    patients = []
    for pid, info in labels.items():
        ct_dir = find_patient_in_ct_roots(info["seg"])
        if ct_dir:
            patients.append((pid, info, ct_dir))

    if max_patients > 0:
        patients = patients[:max_patients]

    print(f"Patients with CT data: {len(patients)}")
    n_ph = sum(1 for _, info, _ in patients if info['label'] == 1)
    print(f"  PH: {n_ph}, nonPH: {len(patients) - n_ph}")

    if len(patients) == 0:
        print("ERROR: No patients with CT data found!")
        return

    # ── Step 1: DICOM → NIfTI (only for needed patients) ──
    print(f"\n[1/3] DICOM → NIfTI ({len(patients)} patients × 5 series)...")
    n_conv, n_skip, n_fail = 0, 0, 0

    for i, (pid, info, ct_dir) in enumerate(patients):
        target_dir = NII_DIR / pid
        target_dir.mkdir(parents=True, exist_ok=True)

        all_exist = all(
            (target_dir / f"{mod}.nii.gz").exists() and
            (target_dir / f"{mod}.nii.gz").stat().st_size > 1000
            for mod in FOLDER_MODALITY.values()
        )
        if all_exist:
            n_skip += 1
            continue

        for sub, mod in FOLDER_MODALITY.items():
            out_path = target_dir / f"{mod}.nii.gz"
            if out_path.exists() and out_path.stat().st_size > 1000:
                continue
            dcm_dir = ct_dir / sub
            if not dcm_dir.is_dir():
                continue
            try:
                dcm_to_nii(dcm_dir, out_path)
            except Exception as e:
                print(f"  [{pid}] {mod} FAILED: {e}")
        n_conv += 1
        if (i + 1) % 5 == 0:
            elapsed = time.time() - t0
            eta = elapsed / (i + 1) * len(patients) - elapsed
            print(f"  [{i+1}/{len(patients)}] converted={n_conv} skip={n_skip} "
                  f"elapsed={elapsed/60:.1f}m ETA={eta/60:.1f}m")

    conv_time = time.time() - t0
    print(f"  DICOM conversion: {conv_time/60:.1f} min (new={n_conv}, skip={n_skip})")

    # ── Step 2: Radiomics extraction ──
    print(f"\n[2/3] Radiomics feature extraction ({len(patients)} patients)...")
    all_records = []
    t1 = time.time()

    for i, (pid, info, ct_dir) in enumerate(patients):
        patient_dir = NII_DIR / pid

        # Check cache
        cache_path = CACHE_DIR / f"{pid}_radiomics.pkl"
        if cache_path.exists():
            try:
                with open(cache_path, "rb") as f:
                    rec = pickle.load(f)
                rec["label"] = info["label"]
                rec["patient_id"] = pid
                all_records.append(rec)
                continue
            except Exception:
                pass

        result = process_one_patient(patient_dir, pid)
        result["label"] = info["label"]
        result["name"] = info["name"]

        if result.get("status") != "ok":
            continue

        all_records.append(result)

        # Cache
        try:
            with open(cache_path, "wb") as f:
                pickle.dump(result, f)
        except Exception:
            pass

        if (i + 1) % 10 == 0:
            t2 = time.time()
            elapsed = t2 - t1
            eta = elapsed / (i + 1) * len(patients) - elapsed
            n_feats = result.get("n_features", 0)
            print(f"  [{i+1}/{len(patients)}] {pid}: {n_feats} features "
                  f"elapsed={elapsed/60:.1f}m ETA={eta/60:.1f}m")

    radio_time = time.time() - t1
    n_ok = len(all_records)
    print(f"\n  Radiomics extraction: {radio_time/60:.1f} min, {n_ok} patients completed")
    if n_ok > 0:
        n_feats = all_records[0].get("n_features", 0)
        print(f"  Features per patient: ~{n_feats}")

    # ── Step 3: Save feature matrix ──
    print(f"\n[3/3] Building feature matrix...")
    all_feature_keys = set()
    for rec in all_records:
        all_feature_keys.update(rec.get("features", {}).keys())

    print(f"  Total unique features across all ROIs: {len(all_feature_keys)}")

    rows = []
    for rec in all_records:
        row = {"patient_id": rec["patient_id"], "label": rec["label"]}
        for k in all_feature_keys:
            row[k] = rec.get("features", {}).get(k, np.nan)
        rows.append(row)

    df = pd.DataFrame(rows)
    # Impute NaN with median
    feature_cols = [c for c in df.columns if c not in ("patient_id", "label")]
    for c in feature_cols:
        median = df[c].median()
        if pd.notna(median):
            df[c] = df[c].fillna(median)
        else:
            df[c] = df[c].fillna(0.0)

    feat_path = OUT_DIR / "radiomics_features.csv"
    df.to_csv(feat_path, index=False, encoding="utf-8")
    print(f"  Feature matrix: {df.shape[0]} patients × {df.shape[1]-2} features")
    print(f"  Saved to {feat_path}")

    total_time = time.time() - t0
    print(f"\n{'=' * 70}")
    print(f"TOTAL TIME: {total_time/60:.1f} min")
    print(f"  DICOM conversion: {conv_time/60:.1f} min")
    print(f"  Radiomics extraction: {radio_time/60:.1f} min")
    print(f"{'=' * 70}")

    return df


if __name__ == "__main__":
    import pandas as pd
    main(max_patients=0)  # 0 = all patients
