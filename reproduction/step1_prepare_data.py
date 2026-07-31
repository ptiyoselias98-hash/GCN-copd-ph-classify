"""
Step 1: Data Preparation — DICOM→NIfTI conversion + label mapping.
==================================================================

1. Parse copd-ph患者113例0331.xlsx → patient_id, name, PH label, mPAP
2. Match patients to CT data directories (00000001=CT, 00000002=Artery,
   00000003=Vein, 00000004=Airway, 00000005=Lung)
3. Validate DICOM count consistency across 00000001-00000005
4. Convert DICOM series → NIfTI (.nii.gz) using SimpleITK
5. Generate labels.csv (patient_id, label, mPAP)
"""
from __future__ import annotations

import json
import os
import re
import shutil
import sys
import tempfile
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import SimpleITK as sitk

# ─── Paths ───────────────────────────────────────────────────────────
REPO_ROOT = Path(__file__).resolve().parent.parent
REPRO_DIR = REPO_ROOT / "reproduction"
NII_ROOT  = REPRO_DIR / "nii_data"
CACHE_DIR = REPRO_DIR / "cache"
OUT_DIR   = REPRO_DIR / "outputs"

EXCEL_PATH = Path(r"E:\桌面文件\资料集-基于大模型与多智能体的COPD-PH资料集\copd-ph患者113例0331.xlsx")

CT_ROOTS = {
    "nonPH": Path(r"H:\官方数据data\COPDnonPH_seg（27例增强性CT)"),
    "PH":    Path(r"H:\官方数据data\COPDPH_seg（160例增强性CT)"),
}

# 00000001 → CT, 00000002 → Artery, 00000003 → Vein, 00000004 → Airway, 00000005 → Lung
FOLDER_TO_MODALITY = {
    "00000001": "ct",
    "00000002": "artery",
    "00000003": "vein",
    "00000004": "airway",
    "00000005": "lung",
}

# ─── Helpers ──────────────────────────────────────────────────────────

def slugify_cn(name: str) -> str:
    """Slugify Chinese patient name to pinyin-like English ID."""
    # Just use a hash-based approach since we don't have pypinyin
    import hashlib
    h = hashlib.md5(name.encode("utf-8")).hexdigest()[:8]
    # Also try to keep readable: replace Chinese chars, keep ASCII
    ascii_part = re.sub(r'[^\x00-\x7F]+', '', name).strip().lower()
    ascii_part = re.sub(r'[^a-z0-9]+', '_', ascii_part).strip('_')
    if ascii_part:
        return ascii_part
    return f"pt_{h}"


def parse_excel(path: Path) -> List[dict]:
    """Parse the wide-format Excel into patient records.

    Excel structure: 424 rows x 114 cols (1 header + 113 patients)
    Row 1 = field names
    Each subsequent row = 1 patient with 424 feature values across columns
    """
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Find key column indices
    col_idx = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        hs = str(h)
        if hs == "PH":
            col_idx["PH"] = i  # 0-indexed
        elif hs == "name":
            col_idx["name"] = i
        elif hs == "mPAP":
            col_idx["mPAP"] = i
        elif hs == "Age":
            col_idx["age"] = i
        elif hs == "Sex":
            col_idx["sex"] = i
        elif "分割" in hs:
            col_idx["seg_folder"] = i
        elif hs == "patient_sn":
            col_idx["patient_sn"] = i

    patients = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=col_idx["name"] + 1).value
        ph_raw = ws.cell(row=row, column=col_idx["PH"] + 1).value
        mpap = ws.cell(row=row, column=col_idx["mPAP"] + 1).value
        age = ws.cell(row=row, column=col_idx.get("age", 0) + 1).value if "age" in col_idx else None
        seg_folder = ws.cell(row=row, column=col_idx.get("seg_folder", 0) + 1).value if "seg_folder" in col_idx else None
        patient_sn = ws.cell(row=row, column=col_idx.get("patient_sn", 0) + 1).value if "patient_sn" in col_idx else None

        # Map label
        if ph_raw == "是":
            label = 1  # COPD-PH
        elif ph_raw == "/":
            label = 0  # COPD-nonPH
        else:
            print(f"  WARNING: unknown PH value '{ph_raw}' for {name}, skipping")
            continue

        # Clean seg_folder (remove trailing /)
        if seg_folder:
            seg_folder = str(seg_folder).rstrip("/").rstrip("\\")

        patients.append({
            "patient_sn": patient_sn,
            "name": name,
            "label": label,
            "mPAP": float(mpap) if mpap is not None else None,
            "age": float(age) if age is not None else None,
            "seg_folder": seg_folder,
            "ph_raw": ph_raw,
        })

    return patients


def find_ct_folder(seg_name: str) -> Optional[Tuple[Path, str]]:
    """Find a patient's CT folder by matching seg_folder name.

    Returns (full_path, label_group) or None.
    """
    seg_clean = str(seg_name).rstrip("/").rstrip("\\")
    for group, root in CT_ROOTS.items():
        candidate = root / seg_clean
        if candidate.is_dir():
            return (candidate, group)
    return None


def validate_dicom_consistency(ct_folder: Path) -> Tuple[bool, Dict[str, int]]:
    """Check that all 5 subfolders have the same number of DICOM files."""
    counts = {}
    for sub in ["00000001", "00000002", "00000003", "00000004", "00000005"]:
        subdir = ct_folder / sub
        if subdir.is_dir():
            dcms = len([f for f in os.listdir(str(subdir))
                       if f.upper().endswith(".DCM")])
            counts[sub] = dcms
        else:
            counts[sub] = 0
    consistent = len(set(counts.values())) <= 1
    return consistent, counts


def convert_dcm_to_nii(dicom_dir: Path, out_path: Path) -> dict:
    """Convert DICOM series to NIfTI using SimpleITK."""
    reader = sitk.ImageSeriesReader()
    series_ids = reader.GetGDCMSeriesIDs(str(dicom_dir))
    if not series_ids:
        return {"status": "no_series"}

    best_files, best_count = None, -1
    for sid in series_ids:
        files = reader.GetGDCMSeriesFileNames(str(dicom_dir), sid)
        if len(files) > best_count:
            best_files, best_count = files, len(files)

    reader.SetFileNames(best_files)
    image = reader.Execute()
    spacing = image.GetSpacing()
    size = image.GetSize()

    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Write via temp file (SimpleITK has issues with long paths on Windows)
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".nii.gz", dir=str(REPRO_DIR))
    os.close(tmp_fd)
    try:
        sitk.WriteImage(image, tmp_path, useCompression=True)
        shutil.move(tmp_path, str(out_path))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    return {
        "status": "ok",
        "n_slices": best_count,
        "shape": list(size),
        "spacing": [float(s) for s in spacing],
    }


def main():
    """Main data preparation pipeline."""
    t0 = time.time()

    print("=" * 70)
    print("STEP 1: Data Preparation — DICOM→NIfTI + Label Mapping")
    print("=" * 70)

    # ── 1. Parse Excel ──
    print("\n[1/5] Parsing Excel clinical data...")
    patients = parse_excel(EXCEL_PATH)
    print(f"  Found {len(patients)} patients in Excel")
    n_ph = sum(1 for p in patients if p["label"] == 1)
    n_nonph = sum(1 for p in patients if p["label"] == 0)
    print(f"  PH={n_ph}, nonPH={n_nonph}")

    # ── 2. Match patients to CT data ──
    print("\n[2/5] Matching patients to CT data folders...")
    matched = []
    missing_ct = []
    for p in patients:
        result = find_ct_folder(p["seg_folder"])
        if result:
            ct_path, group = result
            consistent, counts = validate_dicom_consistency(ct_path)
            p["ct_folder"] = str(ct_path)
            p["ct_group"] = group
            p["dicom_counts"] = counts
            p["dicom_consistent"] = consistent
            matched.append(p)
        else:
            missing_ct.append(p)

    print(f"  Matched: {len(matched)}, Missing CT: {len(missing_ct)}")
    if missing_ct:
        print("  MISSING CT DATA (will be excluded):")
        for p in missing_ct:
            print(f"    - {p['name']}: seg_folder='{p['seg_folder']}'")

    # ── 3. Validate DICOM consistency ──
    print("\n[3/5] Validating DICOM consistency...")
    inconsistent = [p for p in matched if not p["dicom_consistent"]]
    if inconsistent:
        print(f"  WARNING: {len(inconsistent)} patients have inconsistent DICOM counts!")
        for p in inconsistent:
            print(f"    - {p['name']}: {p['dicom_counts']}")
    else:
        print(f"  All {len(matched)} patients have consistent DICOM counts ✓")

    # ── 4. Convert DICOM → NIfTI ──
    print(f"\n[4/5] Converting DICOM → NIfTI for {len(matched)} patients...")
    nii_root = NII_ROOT
    nii_root.mkdir(parents=True, exist_ok=True)

    conversion_log = []
    n_ok, n_skip, n_fail = 0, 0, 0

    for i, p in enumerate(matched):
        ct_folder = Path(p["ct_folder"])
        patient_id = slugify_cn(p["name"])
        p["patient_id"] = patient_id  # store for later use

        target_dir = nii_root / patient_id
        target_dir.mkdir(parents=True, exist_ok=True)

        # Count existing valid nii files
        existing = sum(1 for f in target_dir.glob("*.nii.gz")
                      if f.stat().st_size > 1000)
        if existing >= 5:
            n_skip += 1
            if (i + 1) % 20 == 0:
                print(f"  [{i+1}/{len(matched)}] {patient_id}: all exists, skip")
            continue

        status = {"patient_id": patient_id, "name": p["name"],
                  "label": p["label"], "modalities": {}}

        all_ok = True
        for sub, modality in FOLDER_TO_MODALITY.items():
            dcm_dir = ct_folder / sub
            out_path = target_dir / f"{modality}.nii.gz"

            if out_path.exists() and out_path.stat().st_size > 1000:
                status["modalities"][modality] = "exists"
                continue

            try:
                rec = convert_dcm_to_nii(dcm_dir, out_path)
                status["modalities"][modality] = rec
                if rec.get("status") != "ok":
                    all_ok = False
            except Exception as exc:
                status["modalities"][modality] = {"status": "error", "error": str(exc)}
                all_ok = False

        if all_ok:
            n_ok += 1
        else:
            n_fail += 1

        conversion_log.append(status)
        if (i + 1) % 10 == 0 or i == 0:
            n_done = sum(1 for s in conversion_log
                        if all(v.get("status") in ("ok", "exists")
                              for v in s["modalities"].values()))
            print(f"  [{i+1}/{len(matched)}] converted={n_done}, "
                  f"total_ok={n_ok}, skip={n_skip}, fail={n_fail}")

    print(f"\n  Conversion complete: ok={n_ok}, skip={n_skip}, fail={n_fail}")

    # ── 5. Generate labels.csv ──
    print("\n[5/5] Generating labels.csv and splits...")
    save_labels_and_splits(matched, nii_root)

    # ── Save full manifest ──
    manifest_path = REPRO_DIR / "patient_manifest.json"
    manifest = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "total_in_excel": len(patients),
        "matched_with_ct": len(matched),
        "missing_ct": len(missing_ct),
        "conversion_ok": n_ok,
        "conversion_skip": n_skip,
        "conversion_fail": n_fail,
        "patients": [],
    }
    for p in matched:
        manifest["patients"].append({
            "patient_id": p.get("patient_id", ""),
            "name": p.get("name", ""),
            "label": p["label"],
            "mPAP": p["mPAP"],
            "age": p["age"],
            "dicom_consistent": p["dicom_consistent"],
            "dicom_counts": p.get("dicom_counts", {}),
        })

    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f"  Manifest saved to {manifest_path}")

    elapsed = time.time() - t0
    print(f"\n{'=' * 70}")
    print(f"STEP 1 COMPLETE — {elapsed/60:.1f} min")
    print(f"  Patients: {len(matched)} ready for graph building")
    print(f"  Labels:   {n_ph} PH, {n_nonph} nonPH")
    print(f"{'=' * 70}")


def save_labels_and_splits(matched: List[dict], nii_root: Path):
    """Generate labels.csv and stratified splits."""
    import pandas as pd

    # labels.csv
    rows = []
    for p in matched:
        pid = p.get("patient_id", slugify_cn(p["name"]))
        rows.append({
            "patient_id": pid,
            "label": p["label"],
            "mPAP": p["mPAP"] if p["mPAP"] is not None else "",
            "age": p["age"] if p["age"] is not None else "",
            "name": p.get("name", ""),
        })

    df = pd.DataFrame(rows)
    labels_path = REPRO_DIR / "labels.csv"
    df.to_csv(labels_path, index=False, encoding="utf-8")
    print(f"  labels.csv ({len(df)} patients) → {labels_path}")
    print(f"    PH (label=1): {(df['label']==1).sum()}")
    print(f"    nonPH (label=0): {(df['label']==0).sum()}")

    # Stratified splits (leave-one-group-out style for small n)
    # For now, just create 5-fold stratified indices
    from sklearn.model_selection import StratifiedKFold
    labels_arr = df["label"].values
    skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=42)

    splits = []
    for fold, (train_idx, val_idx) in enumerate(skf.split(np.arange(len(df)), labels_arr)):
        train_pids = df.iloc[train_idx]["patient_id"].tolist()
        val_pids = df.iloc[val_idx]["patient_id"].tolist()
        splits.append({"fold": fold, "train": train_pids, "val": val_pids})

    splits_path = REPRO_DIR / "splits" / "splits_stratified_5fold.json"
    with open(splits_path, "w", encoding="utf-8") as f:
        json.dump(splits, f, indent=2, ensure_ascii=False)
    print(f"  Stratified 5-fold splits → {splits_path}")

    # Also save as separate train/val files per fold
    for s in splits:
        fold_dir = REPRO_DIR / "splits" / f"fold_{s['fold']}"
        fold_dir.mkdir(parents=True, exist_ok=True)
        (fold_dir / "train.txt").write_text("\n".join(s["train"]), encoding="utf-8")
        (fold_dir / "val.txt").write_text("\n".join(s["val"]), encoding="utf-8")


if __name__ == "__main__":
    main()
