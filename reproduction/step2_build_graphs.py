"""
Step 2: Build Graphs — NIfTI masks → PyG Data objects + Radiomics extraction.
=============================================================================

For each patient:
  1. Load NIfTI masks (artery.nii.gz, vein.nii.gz, lung.nii.gz, airway.nii.gz, ct.nii.gz)
  2. Combine artery+vein → vessel mask → 3D skeleton extraction
  3. Trace branches, compute morphological features
  4. Build PyG graph (nodes=bifurcations, edges=segments)
  5. Extract quantitative features (BV5, LAA%, pruning index, etc.)
  6. Extract radiomics from Excel clinical data
  7. Cache as .pkl files

Uses the cloned repo's utils modules (copdph-gcn-repo/utils/).
"""

from __future__ import annotations

import json
import os
import pickle
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional

import nibabel as nib
import numpy as np
import pandas as pd

# Add the repo to path
REPO_ROOT = Path(__file__).resolve().parent.parent
REPRO_DIR = REPO_ROOT / "reproduction"
NII_ROOT  = REPRO_DIR / "nii_data"
CACHE_DIR = REPRO_DIR / "cache"
OUT_DIR   = REPRO_DIR / "outputs"

sys.path.insert(0, str(REPO_ROOT / "copdph-gcn-repo"))

from utils.skeleton import VesselSkeleton
from utils.graph_builder import VascularGraphBuilder
from utils.quantification import (
    VascularQuantifier, ParenchymaQuantifier, AirwayQuantifier,
    extract_all_features,
)
from enhance_features import augment_graph, compute_node_curvature


def load_nifti(path: Path) -> tuple[np.ndarray, tuple]:
    """Load a NIfTI file, return (data_array, spacing)."""
    img = nib.load(str(path))
    arr = np.asarray(img.get_fdata(), dtype=np.float32)
    zooms = img.header.get_zooms()[:3]
    spacing = tuple(float(z) for z in zooms)
    return arr, spacing


def process_one_patient(
    patient_dir: Path,
    case_id: str,
    label: int,
    mpap: Optional[float] = None,
    config: Optional[dict] = None,
) -> Optional[dict]:
    """Process a single patient: NIfTI → graph + features.

    Returns dict with keys: graph, features, label, patient_id, mpap
    """
    if config is None:
        config = {}

    # ── 1. Load all masks ──
    try:
        ct_arr, spacing = load_nifti(patient_dir / "ct.nii.gz")
        artery_arr, _ = load_nifti(patient_dir / "artery.nii.gz")
        vein_arr, _ = load_nifti(patient_dir / "vein.nii.gz")
        lung_arr, _ = load_nifti(patient_dir / "lung.nii.gz")
        airway_arr, _ = load_nifti(patient_dir / "airway.nii.gz")
    except FileNotFoundError as e:
        print(f"  [{case_id}] MISSING FILE: {e}")
        return None

    artery_mask = (artery_arr > 0).astype(np.uint8)
    vein_mask   = (vein_arr > 0).astype(np.uint8)
    lung_mask   = (lung_arr > 0).astype(np.uint8)
    airway_mask = (airway_arr > 0).astype(np.uint8)

    # Combine artery + vein → total vessel mask
    vessel_mask = ((artery_mask + vein_mask) > 0).astype(np.uint8)
    if vessel_mask.sum() < 100:  # too few vessel voxels
        print(f"  [{case_id}] SKIP: vessel mask too small ({vessel_mask.sum()} voxels)")
        return None

    # ── 2. Skeleton extraction ──
    skel_cfg = config.get("skeleton", {})
    skeleton_extractor = VesselSkeleton(
        min_branch_length=skel_cfg.get("min_branch_length", 3),
    )

    try:
        skel = skeleton_extractor.extract_skeleton(vessel_mask)
    except Exception as e:
        print(f"  [{case_id}] SKELETON FAILED: {e}")
        return None

    if skel.sum() < 10:
        print(f"  [{case_id}] SKIP: skeleton too small ({skel.sum()} voxels)")
        return None

    classified = skeleton_extractor.classify_voxels(skel)
    branches = skeleton_extractor.trace_branches(skel, classified)
    if not branches:
        print(f"  [{case_id}] SKIP: no branches traced")
        return None

    # Distance transform for diameter estimation (compute ONCE)
    from scipy import ndimage
    vessel_dt = ndimage.distance_transform_edt(
        vessel_mask > 0, sampling=np.array(spacing)
    )

    branch_features = [
        skeleton_extractor.compute_branch_features(
            b, vessel_mask, ct_volume=ct_arr, spacing=spacing, dt=vessel_dt,
        )
        for b in branches
    ]

    # ── 3. Build PyG graph ──
    gcfg = config.get("graph", {})
    graph_builder = VascularGraphBuilder(
        spatial_edge_threshold=gcfg.get("spatial_edge_threshold", 15.0),
        add_spatial_edges=gcfg.get("add_spatial_edges", True),
        use_directed=gcfg.get("use_directed", False),
    )
    graph = graph_builder.build_graph(branches, branch_features, label=label)

    # ── 4. Quantitative features ──
    qfeat = extract_all_features(
        ct_arr, artery_mask, vein_mask, lung_mask, airway_mask,
        branches, branch_features, spacing,
    )

    # ── 5. Per-structure metrics (artery, vein, airway separately) ──
    vq = VascularQuantifier(spacing)
    pq = ParenchymaQuantifier(spacing)
    aq = AirwayQuantifier(spacing)

    artery_feats = vq.compute_blood_volume_metrics(artery_mask, lung_mask)
    vein_feats   = vq.compute_blood_volume_metrics(vein_mask, lung_mask)
    av_ratio     = vq.compute_artery_vein_ratio(artery_mask, vein_mask, lung_mask)
    parenchyma   = pq.compute_laa(ct_arr, lung_mask)
    airway_feats = aq.compute_airway_metrics(airway_mask, ct_arr, lung_mask)

    # Total vessel volume (needed for diameter calibration in enhance_features)
    voxel_vol = float(np.prod(np.array(spacing)))
    total_vessel_vol_ml = float(np.sum(vessel_mask)) * voxel_vol / 1000.0

    result = {
        "graph": graph,
        "features": {
            "vascular": qfeat.get("vascular", {}),
            "parenchyma": qfeat.get("parenchyma", {}),
            "airway": qfeat.get("airway", {}),
        },
        "quant": {
            "artery": artery_feats,
            "vein": vein_feats,
            "artery_vein_ratio": av_ratio,
            "parenchyma": parenchyma,
            "airway": airway_feats,
            "total_vessel_volume_ml": total_vessel_vol_ml,
            "spacing": list(spacing),
            "num_branches": len(branches),
            "num_nodes": int(graph.num_nodes) if hasattr(graph, 'num_nodes') else 0,
            "num_edges": int(graph.edge_index.size(1)) if hasattr(graph, 'edge_index') else 0,
        },
        "label": int(label),
    }

    if mpap is not None and not (isinstance(mpap, float) and np.isnan(mpap)):
        result["mpap"] = float(mpap)
        result["graph"].mpap = np.array([float(mpap)], dtype=np.float32)

    return result


def build_radiomics_table(
    processed: List[dict],
    excel_path: Path,
) -> pd.DataFrame:
    """Build radiomics feature table from Excel quantitative data.

    Extracts per-patient quantitative CT features (LAA%, vessel volumes,
    density metrics, etc.) and aligns with processed patient IDs.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Column name → index map
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        col_map[str(h)] = i + 1  # 1-indexed

    # Identify radiomics-related columns (quantitative CT features)
    # These are the commercial CT radiomics columns in the Excel
    radio_cols = [
        "全肺容积(ml)", "全肺LAA910(ml)", "全肺LAA950(ml)",
        "全肺LAA910(%)", "全肺LAA950(%)", "全肺平均密度(HU)",
        "全肺肺组织体积(ml)", "全肺空气体积(ml)", "全肺密度标准差(HU)", "全肺质量(g)",
        "右肺容积(ml)_x", "右肺LAA910(ml)_x", "右肺LAA950(ml)_x",
        "右肺LAA910(%)_x", "右肺LAA950(%)_x", "右肺平均密度(HU)_x",
        "右肺肺组织体积(ml)_x", "右肺空气体积(ml)_x", "右肺密度标准差(HU)_x", "右肺质量(g)_x",
        "左肺容积(ml)_x", "左肺LAA910(ml)_x", "左肺LAA950(ml)_x",
        "左肺LAA910(%)_x", "左肺LAA950(%)_x", "左肺平均密度(HU)_x",
    ]

    # Also look for vessel-specific columns
    vessel_keywords = ["血管", "动脉", "静脉", "BV5", "BV10", "分形", "弯曲"]
    for h in headers:
        if h is None:
            continue
        hs = str(h)
        for kw in vessel_keywords:
            if kw in hs:
                if hs not in radio_cols:
                    radio_cols.append(hs)
                break

    print(f"  Radiomics columns found: {len(radio_cols)}")

    # Build patient name → processed patient lookup
    name_to_pid = {}
    for p in processed:
        name_to_pid[p.get("name", "")] = p["patient_id"]

    # Also build seg_folder → patient_id lookup
    seg_to_pid = {}
    for p in processed:
        seg_to_pid[p.get("seg_folder", "")] = p["patient_id"]

    records = []
    name_col = col_map.get("name", 3)
    seg_col = col_map.get("分割文件夹", 12)
    ph_col = col_map.get("PH", 7)
    mpap_col = col_map.get("mPAP", 14)

    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        seg_folder = ws.cell(row=row, column=seg_col).value
        if seg_folder:
            seg_folder = str(seg_folder).rstrip("/").rstrip("\\")

        # Match to processed patient
        pid = seg_to_pid.get(seg_folder, "") or name_to_pid.get(str(name) if name else "", "")

        if not pid:
            continue

        record = {"patient_id": pid, "name": name}
        for cn in radio_cols:
            ci = col_map.get(cn)
            if ci:
                val = ws.cell(row=row, column=ci).value
                try:
                    record[cn] = float(val) if val is not None else np.nan
                except (ValueError, TypeError):
                    record[cn] = np.nan

        records.append(record)

    df = pd.DataFrame(records)
    df = df.set_index("patient_id")

    # Fill NaN with column median
    for c in df.columns:
        if c == "name":
            continue
        median_val = df[c].median()
        if pd.notna(median_val):
            df[c] = df[c].fillna(median_val)

    return df


def main():
    t0 = time.time()
    print("=" * 70)
    print("STEP 2: Graph Building + Radiomics Extraction")
    print("=" * 70)

    # ── Load labels ──
    labels_path = REPRO_DIR / "labels.csv"
    if not labels_path.exists():
        print("ERROR: labels.csv not found. Run step1_prepare_data.py first.")
        return 1

    labels_df = pd.read_csv(labels_path)
    print(f"\n[1/4] Loaded {len(labels_df)} patients from labels.csv")
    print(f"  PH: {(labels_df['label']==1).sum()}, nonPH: {(labels_df['label']==0).sum()}")

    # ── Load patient manifest for seg_folder info ──
    manifest_path = REPRO_DIR / "patient_manifest.json"
    seg_info = {}
    if manifest_path.exists():
        with open(manifest_path, "r", encoding="utf-8") as f:
            manifest = json.load(f)
        for p in manifest.get("patients", []):
            pid = p.get("patient_id", "")
            if pid:
                seg_info[pid] = {
                    "name": p.get("name", ""),
                    "seg_folder": p.get("dicom_counts", {}),  # not ideal, need actual seg info
                }
    else:
        print("  WARNING: patient_manifest.json not found, will use labels.csv only")

    # ── Process each patient ──
    print(f"\n[2/4] Building graphs for {len(labels_df)} patients...")
    os.makedirs(CACHE_DIR, exist_ok=True)

    processed = []
    n_ok, n_skip, n_fail = 0, 0, 0

    config = {
        "skeleton": {"min_branch_length": 3},
        "graph": {
            "spatial_edge_threshold": 15.0,
            "add_spatial_edges": True,
            "use_directed": False,
        },
    }

    for idx, row in labels_df.iterrows():
        pid = row["patient_id"]
        label = int(row["label"])
        mpap = row.get("mPAP", None)
        try:
            mpap = float(mpap) if pd.notna(mpap) else None
        except (ValueError, TypeError):
            mpap = None

        patient_dir = NII_ROOT / pid
        cache_path = CACHE_DIR / f"{pid}.pkl"

        if not patient_dir.is_dir():
            print(f"  [{pid}] MISSING NIfTI directory, skipping")
            n_skip += 1
            continue

        # Check cache
        if cache_path.exists():
            try:
                with open(cache_path, "rb") as f:
                    entry = pickle.load(f)
                entry["patient_id"] = pid
                entry["label"] = label
                processed.append(entry)
                n_ok += 1
                if (n_ok % 20) == 0:
                    print(f"  [{n_ok}/{len(labels_df)}] cache hit, skipped processing")
                continue
            except Exception:
                pass  # corrupted cache, rebuild

        # Process
        result = process_one_patient(patient_dir, pid, label, mpap, config)
        if result is None:
            n_fail += 1
            print(f"  [{pid}] FAILED: no valid graph produced")
            continue

        # Save cache (without patient_id to match upstream convention)
        try:
            cache_data = {k: v for k, v in result.items() if k != "patient_id"}
            with open(cache_path, "wb") as f:
                pickle.dump(cache_data, f)
        except Exception as e:
            print(f"  [{pid}] CACHE WRITE FAILED: {e}")

        result["patient_id"] = pid
        processed.append(result)
        n_ok += 1

        if (n_ok % 10) == 0:
            nnodes = result["quant"]["num_nodes"]
            print(f"  [{n_ok}/{len(labels_df)}] {pid}: "
                  f"nodes={nnodes}, branches={result['quant']['num_branches']}")

    # Summary
    print(f"\n  Graph building: ok={n_ok}, skip={n_skip}, fail={n_fail}")
    if n_ok == 0:
        print("ERROR: No valid graphs produced!")
        return 1

    # ── Extract radiomics ──
    print(f"\n[3/4] Extracting radiomics features...")
    excel_path = Path(r"E:\桌面文件\资料集-基于大模型与多智能体的COPD-PH资料集\copd-ph患者113例0331.xlsx")
    radiomics_df = build_radiomics_table(processed, excel_path)

    if len(radiomics_df) == 0:
        print("  WARNING: No radiomics could be extracted from Excel!")
        # Build fallback radiomics from in-pipeline quantification
        print("  Building fallback radiomics from pipeline quantification...")
        rad_records = []
        for p in processed:
            q = p.get("quant", {})
            rad_records.append({
                "patient_id": p["patient_id"],
                **{f"q_{k}": v for k, v in q.items()
                   if isinstance(v, (int, float)) and not isinstance(v, bool)},
            })
        radiomics_df = pd.DataFrame(rad_records).set_index("patient_id")

    # Save radiomics
    radiomics_path = REPRO_DIR / "radiomics.csv"
    radiomics_df.to_csv(radiomics_path, encoding="utf-8")
    print(f"  Radiomics saved: {radiomics_df.shape[0]} patients × {radiomics_df.shape[1]} features")
    print(f"  → {radiomics_path}")

    # ── Build enhanced graphs with global features ──
    print(f"\n[4/4] Enhancing graphs with commercial scalars...")
    n_enhanced = 0
    for p in processed:
        pid = p["patient_id"]
        graph = p["graph"]

        # Get commercial scalars from radiomics if available
        if pid in radiomics_df.index:
            row = radiomics_df.loc[pid]
            try:
                commercial_total_vol = float(row.get("全肺容积(ml)", np.nan))
                commercial_laa910 = float(row.get("全肺LAA910(%)", np.nan))
                commercial_laa950 = float(row.get("全肺LAA950(%)", np.nan))
                commercial_lung_density = float(row.get("全肺平均密度(HU)", np.nan))
                commercial_lung_std = float(row.get("全肺密度标准差(HU)", np.nan))

                # Build enhanced graph
                enhanced = augment_graph(
                    graph,
                    commercial_total_vol_ml=commercial_total_vol if np.isfinite(commercial_total_vol) else None,
                    commercial_fractal_dim=None,
                    commercial_artery_density=None,
                    commercial_vein_density=None,
                    pipeline_total_vol_ml=p["quant"].get("total_vessel_volume_ml", None),
                    commercial_lung_density_std=commercial_lung_std if np.isfinite(commercial_lung_std) else None,
                )
                p["graph_enhanced"] = enhanced
                n_enhanced += 1
            except Exception as e:
                p["graph_enhanced"] = graph  # fallback to baseline
        else:
            p["graph_enhanced"] = graph

    print(f"  Enhanced graphs: {n_enhanced}/{len(processed)}")

    # ── Save processed dataset manifest ──
    proc_manifest = []
    for p in processed:
        proc_manifest.append({
            "patient_id": p["patient_id"],
            "label": p["label"],
            "mpap": p.get("mpap"),
            "num_nodes": p["quant"]["num_nodes"],
            "num_edges": p["quant"]["num_edges"],
            "num_branches": p["quant"]["num_branches"],
        })

    proc_path = REPRO_DIR / "processed_manifest.json"
    with open(proc_path, "w", encoding="utf-8") as f:
        json.dump(proc_manifest, f, indent=2, ensure_ascii=False)

    # Stats
    labels_list = [p["label"] for p in processed]
    n_ph = sum(labels_list)
    n_nonph = len(labels_list) - n_ph
    nnodes_list = [p["quant"]["num_nodes"] for p in processed]
    nbranches_list = [p["quant"]["num_branches"] for p in processed]

    elapsed = time.time() - t0
    print(f"\n{'=' * 70}")
    print(f"STEP 2 COMPLETE — {elapsed/60:.1f} min")
    print(f"  Graphs built: {n_ok}")
    print(f"  PH: {n_ph}, nonPH: {n_nonph}")
    print(f"  Nodes: mean={np.mean(nnodes_list):.0f}, "
          f"median={np.median(nnodes_list):.0f}, "
          f"range=[{min(nnodes_list)}, {max(nnodes_list)}]")
    print(f"  Branches: mean={np.mean(nbranches_list):.0f}, "
          f"median={np.median(nbranches_list):.0f}")
    print(f"{'=' * 70}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
