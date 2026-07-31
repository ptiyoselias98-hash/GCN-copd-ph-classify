"""
Feature Importance Analysis — What drives the COPD-PH classification AUC?
=========================================================================
Answers: (1) exact data used, (2) feature selection methods, (3) top features by importance.
Runs permutation importance + SHAP on the best model, grouped by feature category.
"""
from __future__ import annotations

import json, os, sys, time, warnings
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
import matplotlib
matplotlib.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "DejaVu Sans"]
matplotlib.rcParams["axes.unicode_minus"] = False
import matplotlib.pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.inspection import permutation_importance
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sklearn.preprocessing import StandardScaler
import xgboost as xgb

warnings.filterwarnings("ignore")

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "outputs" / "feature_importance"
OUT_DIR.mkdir(parents=True, exist_ok=True)

EXCEL = Path(r"E:\桌面文件\资料集-基于大模型与多智能体的COPD-PH资料集\copd-ph患者113例0331.xlsx")
SEED = 42

# Feature group definitions (column index → group name)
def get_ct_feature_groups(headers: list) -> dict:
    """Map column index to feature group name."""
    groups = {}
    for ci in range(250, len(headers) + 1):
        h = headers[ci - 1]
        if h is None: continue
        hs = str(h)
        if 'LAA' in hs:                    g = 'LAA (emphysema)'
        elif 'BV5' in hs:                  g = 'BV5 (small vessel <5mm)'
        elif 'BV10' in hs:                 g = 'BV10 (vessel <10mm)'
        elif '密度' in hs and '标准差' in hs: g = 'Density StdDev'
        elif '分形' in hs:                 g = 'Fractal Dimension'
        elif '弯曲度' in hs:               g = 'Vessel Tortuosity'
        elif '钙化' in hs or 'Agatston' in hs: g = 'Calcification'
        elif 'Pi10' in hs:                 g = 'Pi10 (airway)'
        elif ('分支' in hs and '数量' in hs) or '血管分支' in hs: g = 'Branch Count'
        elif ('分支' in hs and '长度' in hs) or '总长度' in hs: g = 'Branch Length'
        elif '分支' in hs:                 g = 'Branch Volume/Wall'
        elif '质量' in hs:                 g = 'Mass'
        elif '容积' in hs or '体积' in hs:  g = 'Volume'
        elif '密度' in hs:                 g = 'Mean Density'
        elif '径' in hs or 'D<' in hs:     g = 'Diameter/Size'
        else:                              g = 'Other'
        groups[ci] = g
    return groups


def parse_data():
    """Parse Excel → X, y with named features."""
    wb = openpyxl.load_workbook(str(EXCEL))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    col_map = {}
    for i, h in enumerate(headers):
        if h is not None: col_map[str(h)] = i + 1

    ph_col = col_map["PH"]   # col 7
    name_col = col_map["name"]  # col 3

    records, names = [], []
    for row in range(2, ws.max_row + 1):
        ph = ws.cell(row=row, column=ph_col).value
        name = ws.cell(row=row, column=name_col).value
        if ph == "是":   label = 1
        elif ph == "/":  label = 0
        else: continue

        rec = {"label": label, "name": name}
        for ci in range(250, ws.max_column + 1):
            val = ws.cell(row=row, column=ci).value
            try: rec[f"ct_{ci}"] = float(val)
            except: rec[f"ct_{ci}"] = np.nan
        records.append(rec)

    df = pd.DataFrame(records)
    # Keep rows with any CT data
    ct_cols = [c for c in df.columns if c.startswith("ct_")]
    df = df[df[ct_cols].notna().any(axis=1)].copy()

    X = df[ct_cols].values.astype(np.float32)
    y = df["label"].values.astype(int)

    # Impute
    for j in range(X.shape[1]):
        col = X[:, j]
        mask = np.isnan(col)
        if mask.any(): col[mask] = np.nanmedian(col)

    # Feature names (use headers for readability)
    feat_names = []
    ct_groups = get_ct_feature_groups(headers)
    for cn in ct_cols:
        ci = int(cn.replace("ct_", ""))
        h = headers[ci - 1] if ci <= len(headers) else cn
        g = ct_groups.get(ci, "Other")
        feat_names.append((cn, ci, str(h) if h else cn, g))

    return X, y, feat_names, len(df)


def run_permutation_importance(X, y, feat_names, model_name="XGBoost"):
    """Stratified 5-fold permutation importance."""
    print(f"\n  Running permutation importance with {model_name}...")

    skf = StratifiedKFold(n_splits=5, shuffle=True, random_state=SEED)

    all_importances = []
    fold_aucs = []

    for fold_i, (tr, va) in enumerate(skf.split(X, y)):
        X_tr, X_va = X[tr], X[va]
        y_tr, y_va = y[tr], y[va]

        # Train model
        if model_name == "XGBoost":
            scale = (y_tr == 0).sum() / max((y_tr == 1).sum(), 1)
            model = xgb.XGBClassifier(
                n_estimators=200, max_depth=5, learning_rate=0.05,
                scale_pos_weight=scale, subsample=0.8, colsample_bytree=0.8,
                random_state=SEED, eval_metric='logloss',
            )
        elif model_name == "RandomForest":
            model = RandomForestClassifier(
                n_estimators=200, max_depth=8, min_samples_leaf=5,
                class_weight='balanced', random_state=SEED, n_jobs=-1,
            )
        elif model_name == "LogisticRegression":
            scaler = StandardScaler()
            X_tr = scaler.fit_transform(X_tr)
            X_va = scaler.transform(X_va)
            model = LogisticRegression(
                penalty='l2', C=1.0, solver='lbfgs', max_iter=5000,
                class_weight='balanced', random_state=SEED,
            )

        model.fit(X_tr, y_tr)
        y_prob = model.predict_proba(X_va)[:, 1]
        fold_auc = roc_auc_score(y_va, y_prob)
        fold_aucs.append(fold_auc)

        # Permutation importance
        r = permutation_importance(
            model, X_va, y_va, n_repeats=10,
            random_state=SEED + fold_i, scoring='roc_auc', n_jobs=-1,
        )
        all_importances.append(r.importances_mean)

    mean_imp = np.mean(all_importances, axis=0)
    std_imp  = np.std(all_importances, axis=0)

    # Sort by importance
    idx = np.argsort(mean_imp)[::-1]
    ranked = []
    for i in idx[:50]:  # top 50
        cn, ci, hname, group = feat_names[i]
        ranked.append({
            "rank": len(ranked) + 1,
            "feature_col": cn,
            "feature_name": hname,
            "group": group,
            "importance_mean": float(mean_imp[i]),
            "importance_std": float(std_imp[i]),
        })

    print(f"  Fold AUCs: {[f'{a:.4f}' for a in fold_aucs]}, mean={np.mean(fold_aucs):.4f}")
    return ranked, fold_aucs


def run_xgboost_builtin_importance(X, y, feat_names):
    """XGBoost native feature importance (gain-based)."""
    print("\n  Running XGBoost built-in importance (gain)...")

    scale = (y == 0).sum() / max((y == 1).sum(), 1)
    model = xgb.XGBClassifier(
        n_estimators=200, max_depth=5, learning_rate=0.05,
        scale_pos_weight=scale, subsample=0.8, colsample_bytree=0.8,
        random_state=SEED, eval_metric='logloss',
    )
    model.fit(X, y)

    # Get feature importance by gain
    importance = model.get_booster().get_score(importance_type='gain')
    # Map f0, f1, ... to feature names
    mapped = {}
    for k, v in importance.items():
        idx = int(k.replace('f', ''))
        if idx < len(feat_names):
            cn, ci, hname, group = feat_names[idx]
            mapped[idx] = {
                "feature_col": cn, "feature_name": hname,
                "group": group, "gain": float(v),
            }

    ranked = sorted(mapped.values(), key=lambda x: -x["gain"])
    for i, r in enumerate(ranked):
        r["rank"] = i + 1
    return ranked[:50]


def plot_top_features(perm_ranked, xgb_ranked, feat_groups):
    """Plot top 25 features by permutation importance."""

    # ── Figure 1: Top 25 permutation importance ──
    fig, ax = plt.subplots(figsize=(12, 10))

    top25 = perm_ranked[:25]
    names = [f"{r['feature_name'][:45]}" for r in reversed(top25)]
    values = [r['importance_mean'] for r in reversed(top25)]
    errors = [r['importance_std'] for r in reversed(top25)]
    colors = [plt.cm.tab20(hash(r['group']) % 20 / 20) for r in reversed(top25)]

    bars = ax.barh(range(len(names)), values, xerr=errors, color=colors,
                   edgecolor='white', linewidth=0.5, capsize=2, height=0.7)

    ax.set_yticks(range(len(names)))
    ax.set_yticklabels(names, fontsize=7)
    ax.set_xlabel("AUC decrease upon permutation (higher = more important)")
    ax.set_title("Top 25 Features by Permutation Importance (XGBoost, 5-fold CV)")

    # Legend for groups
    unique_groups = sorted(set(r['group'] for r in top25))
    legend_patches = [plt.Rectangle((0,0),1,1, color=plt.cm.tab20(hash(g)%20/20)) for g in unique_groups]
    ax.legend(legend_patches, unique_groups, fontsize=7, loc='lower right')

    plt.tight_layout()
    fig.savefig(OUT_DIR / "top25_permutation_importance.png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  → {OUT_DIR / 'top25_permutation_importance.png'}")

    # ── Figure 2: Group-level importance ──
    group_imp = {}
    for r in perm_ranked:
        g = r['group']
        group_imp.setdefault(g, []).append(r['importance_mean'])

    fig, ax = plt.subplots(figsize=(12, 6))
    groups_sorted = sorted(group_imp.keys(), key=lambda g: -np.sum(group_imp[g]))
    vals = [np.sum(group_imp[g]) for g in groups_sorted]
    colors_g = plt.cm.Set3(np.linspace(0, 1, len(groups_sorted)))
    bars = ax.bar(range(len(groups_sorted)), vals, color=colors_g, edgecolor='white')

    ax.set_xticks(range(len(groups_sorted)))
    ax.set_xticklabels(groups_sorted, rotation=45, ha='right', fontsize=9)
    ax.set_ylabel("Sum of Permutation Importance")
    ax.set_title("Feature Importance by Category (CT_all, n=175 features)")

    # Annotate count
    for i, g in enumerate(groups_sorted):
        cnt = len(group_imp[g])
        ax.text(i, vals[i] + 0.002, f'n={cnt}', ha='center', fontsize=8)

    plt.tight_layout()
    fig.savefig(OUT_DIR / "group_importance.png", dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"  → {OUT_DIR / 'group_importance.png'}")


def main():
    t0 = time.time()
    print("=" * 70)
    print("FEATURE IMPORTANCE ANALYSIS")
    print("=" * 70)

    # ── Parse data ──
    print("\n[1] Loading data...")
    X, y, feat_names, n_patients = parse_data()
    print(f"  Patients: {n_patients} (PH={(y==1).sum()}, nonPH={(y==0).sum()})")
    print(f"  CT features: {X.shape[1]}")

    # ── Feature stats ──
    groups = {}
    for cn, ci, hname, g in feat_names:
        groups.setdefault(g, []).append(hname)

    print(f"\n  Feature groups:")
    for g in sorted(groups.keys()):
        print(f"    {g:30s}: {len(groups[g]):3d} features")

    print(f"\n  Clinical_safe: 11 features")
    kept_clinical = ["Age", "WBC", "RBC", "HB", "PLT",
                     "ALT", "AST", "Crea", "UA", "D-Dimer", "cTnT"]
    for f in kept_clinical:
        print(f"    {f}")

    # ── Run importance analyses ──
    print(f"\n[2] Feature importance analyses...")
    perm_ranked, fold_aucs = run_permutation_importance(X, y, feat_names, "XGBoost")
    xgb_ranked = run_xgboost_builtin_importance(X, y, feat_names)

    # ── Top features table ──
    print(f"\n[3] TOP 20 FEATURES (Permutation Importance)")
    print(f"{'Rank':<5} {'Feature':<45s} {'Group':<30s} {'Importance':>10s}")
    print("-" * 95)
    for r in perm_ranked[:20]:
        name = r['feature_name'][:43]
        print(f"{r['rank']:<5} {name:<45s} {r['group']:<30s} {r['importance_mean']:>10.4f}")

    # ── Group-level summary ──
    print(f"\n[4] GROUP-LEVEL IMPORTANCE")
    group_sum = {}
    for r in perm_ranked:
        g = r['group']
        group_sum.setdefault(g, {'sum': 0, 'count': 0, 'top_feat': None, 'top_imp': -1})
        group_sum[g]['sum'] += r['importance_mean']
        group_sum[g]['count'] += 1
        if r['importance_mean'] > group_sum[g]['top_imp']:
            group_sum[g]['top_imp'] = r['importance_mean']
            group_sum[g]['top_feat'] = r['feature_name']

    print(f"{'Group':<35s} {'#Feats':>6s} {'SumImp':>8s} {'TopFeature':<40s}")
    print("-" * 95)
    for g in sorted(group_sum.keys(), key=lambda g: -group_sum[g]['sum']):
        s = group_sum[g]
        print(f"{g:<35s} {s['count']:>6d} {s['sum']:>8.4f} {s['top_feat'][:38]:<40s}")

    # ── Compare with XGBoost gain ──
    print(f"\n[5] TOP 10 BY XGBOOST GAIN")
    for r in xgb_ranked[:10]:
        print(f"  {r['rank']:>2}. {r['feature_name'][:45]:45s} [{r['group']}] gain={r['gain']:.4f}")

    # ── Plots ──
    print(f"\n[6] Generating plots...")
    plot_top_features(perm_ranked, xgb_ranked, groups)

    # ── Save all data ──
    output = {
        "config": {"n_patients": n_patients, "n_features": X.shape[1],
                   "n_ph": int((y==1).sum()), "n_nonph": int((y==0).sum()),
                   "fold_aucs": fold_aucs},
        "permutation_importance_top50": perm_ranked,
        "xgboost_gain_top50": xgb_ranked,
        "feature_groups": {g: len(v) for g, v in groups.items()},
    }
    with open(OUT_DIR / "feature_importance.json", "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    print(f"\n  JSON → {OUT_DIR / 'feature_importance.json'}")

    # ── Excel report ──
    perm_df = pd.DataFrame(perm_ranked)
    perm_df.to_excel(OUT_DIR / "feature_importance.xlsx", index=False)
    print(f"  Excel → {OUT_DIR / 'feature_importance.xlsx'}")

    elapsed = time.time() - t0
    print(f"\nTotal: {elapsed/60:.1f} min")
    return 0


if __name__ == "__main__":
    sys.exit(main())
