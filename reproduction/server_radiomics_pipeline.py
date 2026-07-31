#!/usr/bin/env python3
"""
COPD-PH Radiomics Pipeline — Ruan 2025 Method (Server Version)
===============================================================
Run on IMSS GPU server: /home/imss/cw/
Requires: pyradiomics, SimpleITK, scikit-learn, xgboost, pandas, openpyxl

Features extracted per patient (5 ROIs × ~200 features each):
  - Vessel ROI (artery+vein mask): morphology + texture from contrast-enhanced CT
  - Artery-only, Vein-only, Lung, Airway masks

Method matches Ruan et al. 2025 JAHA:
  1. PyRadiomics from original + wavelet + LoG filtered CT images
  2. MWU pre-filter → LASSO feature selection
  3. LRC, RFC, SVC models — Radiomic Model (RM) + Joint Model (JM)
  4. 10-repeat 5-fold CV with 95% CI

Estimated runtime: 40-80 min for 100 patients
"""

import os, sys, time, json, pickle, hashlib, re, warnings
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
import SimpleITK as sitk

warnings.filterwarnings("ignore")

# ─── Paths (SERVER) ────────────────────────────────────────────────
BASE = Path("/home/imss/cw")
NONPH_DIR = BASE / "COPDnonPH COPD-PH"
DATA_ROOT = BASE  # adjust if needed
NII_DIR   = BASE / "radiomics_reproduction" / "nii_data"
CACHE_DIR = BASE / "radiomics_reproduction" / "cache"
OUT_DIR   = BASE / "radiomics_reproduction" / "outputs"
for d in [NII_DIR, CACHE_DIR, OUT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# Excel labels — copy from Windows to server first
EXCEL_PATH = BASE / "copd-ph患者113例0331.xlsx"
LABELS_CSV = BASE / "radiomics_reproduction" / "labels.csv"

FOLDER_MODALITY = {
    "00000001": "ct", "00000002": "artery",
    "00000003": "vein", "00000004": "airway", "00000005": "lung",
}

SEED = 2020
N_REPEATS = 10
N_FOLDS = 5


# ═══════════════════════════════════════════════════════════════
# 1. Labels
# ═══════════════════════════════════════════════════════════════

def build_labels():
    """Parse Excel → labels dict + find matching CT directories."""
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel not found at {EXCEL_PATH}")
        print("Copy copd-ph患者113例0331.xlsx to server first!")
        sys.exit(1)

    import openpyxl
    wb = openpyxl.load_workbook(str(EXCEL_PATH))
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    ph_col = next((i+1 for i,h in enumerate(headers) if h == "PH"), 7)
    name_col = next((i+1 for i,h in enumerate(headers) if h == "name"), 3)
    seg_col = next((i+1 for i,h in enumerate(headers) if h and "分割" in str(h)), 12)

    labels = []
    for row in range(2, ws.max_row + 1):
        ph = ws.cell(row=row, column=ph_col).value
        name = ws.cell(row=row, column=name_col).value
        seg = ws.cell(row=row, column=seg_col).value
        if ph == "是":   label = 1
        elif ph == "/":  label = 0
        else: continue

        seg_clean = str(seg).rstrip("/").rstrip("\\") if seg else ""
        pid = re.sub(r'[^\x00-\x7F]+', '', str(name) if name else '').strip().lower()
        pid = re.sub(r'[^a-z0-9]+', '_', pid).strip('_')
        if not pid:
            h = hashlib.md5(str(name).encode('utf-8')).hexdigest()[:8]
            pid = f'pt_{h}'

        labels.append({"patient_id": pid, "label": label, "name": name, "seg": seg_clean})

    df = pd.DataFrame(labels)
    df.to_csv(LABELS_CSV, index=False)
    print(f"Labels: {len(df)} patients (PH={(df.label==1).sum()}, nonPH={(df.label==0).sum()})")
    return labels


# ═══════════════════════════════════════════════════════════════
# 2. Find CT data
# ═══════════════════════════════════════════════════════════════

def find_ct_dirs(labels: list) -> list:
    """Scan known locations on server for patient CT directories."""
    # Try common locations
    search_paths = [
        Path("/home/imss/cw"),
        Path("/home/imss/data"),
        Path("/home/imss/COPDnonPH COPD-PH"),
    ]

    # Build seg_folder → path index
    seg_to_path = {}
    known_groups = {
        "COPDnonPH_seg（27例增强性CT)": "H:/",
        "COPDPH_seg（160例增强性CT)": "H:/",
    }

    # Actually on the server, find the actual data paths
    # Check known mount points
    possible_roots = [
        Path("/mnt/H/官方数据data/COPDnonPH_seg（27例增强性CT)"),
        Path("/mnt/H/官方数据data/COPDPH_seg（160例增强性CT)"),
        Path("/home/imss/data/COPDnonPH_seg"),
        Path("/home/imss/data/COPDPH_seg"),
        Path("/mnt/data/COPDnonPH"),
        Path("/mnt/data/COPDPH"),
    ]

    # Also search under CW
    cw = Path("/home/imss/cw")
    if cw.exists():
        for d in cw.iterdir():
            if d.is_dir() and ("COPD" in d.name or "PH" in d.name or "seg" in d.name):
                possible_roots.append(d)

    # Print what we found
    print("\nSearching for CT data directories...")
    found_roots = [r for r in possible_roots if r.exists()]
    if not found_roots:
        print("WARNING: No standard CT data paths found!")
        print("Please run: find /home/imss -name '00000001' -type d 2>/dev/null | head -10")
        print("Then update possible_roots in this script.")

    for r in found_roots:
        count = len(list(r.iterdir())) if r.exists() else 0
        print(f"  {r}: {count} subdirectories")

    # Match patients to CT folders
    matched = []
    for p in labels:
        seg = p["seg"]
        found = False
        for root in found_roots:
            candidate = root / seg
            if candidate.is_dir() and (candidate / "00000001").is_dir():
                p["ct_dir"] = str(candidate)
                matched.append(p)
                found = True
                break
        if not found:
            # Try searching by patient name
            for root in found_roots:
                for d in root.iterdir():
                    if d.is_dir() and p["name"] and (p["name"] in d.name or seg[:20] in d.name):
                        if (d / "00000001").is_dir():
                            p["ct_dir"] = str(d)
                            matched.append(p)
                            found = True
                            break
                if found:
                    break

    print(f"Matched: {len(matched)}/{len(labels)} patients with CT data")
    if len(matched) < len(labels):
        missing = [p for p in labels if "ct_dir" not in p]
        print(f"Missing {len(missing)} patients:")
        for m in missing[:5]:
            print(f"  {m['name']}: seg={m['seg'][:40]}")

    return matched


# ═══════════════════════════════════════════════════════════════
# 3. DICOM → NIfTI
# ═══════════════════════════════════════════════════════════════

def dcm_to_nii(dicom_dir: Path, out_path: Path) -> bool:
    """Convert one DICOM series to NIfTI. Returns True on success."""
    try:
        reader = sitk.ImageSeriesReader()
        sids = reader.GetGDCMSeriesIDs(str(dicom_dir))
        if not sids:
            return False
        files = reader.GetGDCMSeriesFileNames(str(dicom_dir), sids[0])
        reader.SetFileNames(files)
        img = reader.Execute()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        sitk.WriteImage(img, str(out_path), useCompression=True)
        return True
    except Exception as e:
        print(f"    FAILED {out_path.name}: {e}")
        return False


def convert_all(patients: list) -> int:
    """Convert DICOM → NIfTI for matched patients."""
    n_new, n_exist = 0, 0
    for i, p in enumerate(patients):
        pid = p["patient_id"]
        ct_dir = Path(p["ct_dir"])
        target_dir = NII_DIR / pid
        target_dir.mkdir(parents=True, exist_ok=True)

        all_exist = True
        for sub, mod in FOLDER_MODALITY.items():
            out = target_dir / f"{mod}.nii.gz"
            if not (out.exists() and out.stat().st_size > 1000):
                all_exist = False
                break

        if all_exist:
            n_exist += 1
            continue

        for sub, mod in FOLDER_MODALITY.items():
            out = target_dir / f"{mod}.nii.gz"
            if out.exists() and out.stat().st_size > 1000:
                continue
            dcm_dir = ct_dir / sub
            if dcm_dir.is_dir():
                dcm_to_nii(dcm_dir, out)

        n_new += 1
        if (i + 1) % 10 == 0:
            print(f"  [{i+1}/{len(patients)}] new={n_new} exist={n_exist}")

    print(f"  DICOM conv: {n_new} converted, {n_exist} existed")
    return n_new + n_exist


# ═══════════════════════════════════════════════════════════════
# 4. PyRadiomics Extraction
# ═══════════════════════════════════════════════════════════════

def extract_radiomics_one(ct_path: Path, mask_path: Path, roi_name: str) -> dict:
    """PyRadiomics feature extraction for one ROI."""
    import radiomics
    from radiomics import featureextractor

    # Initialize extractor with PyRadiomics settings (matching Ruan paper)
    settings = {
        'binWidth': 5,
        'resampledPixelSpacing': [1, 1, 1],
        'interpolator': 'sitkBSpline',
        'verbose': False,
        'enableCExtensions': True,
    }
    extractor = featureextractor.RadiomicsFeatureExtractor(**settings)

    # Enable all feature classes
    extractor.enableAllFeatures()
    extractor.enableImageTypes(Original={}, Wavelet={}, LoG={'sigma': [3.0, 5.0]})

    try:
        result = extractor.execute(str(ct_path), str(mask_path))
        # Filter: keep only diagnostic features, drop general info
        feats = {}
        for k, v in result.items():
            if k.startswith('diagnostics_') or k.startswith('general_'):
                continue
            try:
                feats[f"{roi_name}_{k}"] = float(v)
            except (ValueError, TypeError):
                pass
        return feats
    except Exception as e:
        print(f"    PyRadiomics error for {roi_name}: {e}")
        return {}


def extract_all_radiomics(patients: list) -> pd.DataFrame:
    """Extract PyRadiomics features for all patients."""
    print("\nInstalling/checking pyradiomics...")
    import radiomics
    print(f"  pyradiomics version: {radiomics.__version__}")

    records = []
    t0 = time.time()

    for i, p in enumerate(patients):
        pid = p["patient_id"]
        patient_dir = NII_DIR / pid

        # Check cache
        cache_path = CACHE_DIR / f"{pid}_pyradiomics.json"
        if cache_path.exists():
            try:
                with open(cache_path, "r") as f:
                    feats = json.load(f)
                feats["patient_id"] = pid
                feats["label"] = p["label"]
                records.append(feats)
                continue
            except Exception:
                pass

        ct_file = patient_dir / "ct.nii.gz"
        if not ct_file.exists():
            continue

        all_feats = {"patient_id": pid, "label": p["label"]}

        # Extract for each ROI
        for mod in ["artery", "vein", "lung", "airway"]:
            mask_file = patient_dir / f"{mod}.nii.gz"
            if not mask_file.exists():
                continue
            feats = extract_radiomics_one(ct_file, mask_file, mod)
            all_feats.update(feats)

        # Vessel = artery + vein
        art_file = patient_dir / "artery.nii.gz"
        vei_file = patient_dir / "vein.nii.gz"
        if art_file.exists() and vei_file.exists():
            # Create combined mask on the fly
            art_img = sitk.ReadImage(str(art_file))
            vei_img = sitk.ReadImage(str(vei_file))
            vessel_img = sitk.Or(art_img > 0, vei_img > 0)
            vessel_img = sitk.Cast(vessel_img, sitk.sitkUInt8)
            vessel_path = CACHE_DIR / f"{pid}_vessel_mask.nii.gz"
            sitk.WriteImage(vessel_img, str(vessel_path))
            feats = extract_radiomics_one(ct_file, vessel_path, "vessel")
            all_feats.update(feats)
            vessel_path.unlink(missing_ok=True)

        # Cache
        with open(cache_path, "w") as f:
            json.dump(all_feats, f)

        records.append(all_feats)

        if (i + 1) % 10 == 0 or i == 0:
            elapsed = time.time() - t0
            eta = elapsed / (i + 1) * len(patients) - elapsed
            nf = len(all_feats) - 2
            print(f"  [{i+1}/{len(patients)}] {pid}: ~{nf} feats, "
                  f"elapsed={elapsed/60:.1f}m, ETA={eta/60:.1f}m")

    total_time = time.time() - t0
    print(f"\n  Total extraction time: {total_time/60:.1f} min")

    # Build DataFrame
    df = pd.DataFrame(records)
    # Impute NaN
    feat_cols = [c for c in df.columns if c not in ("patient_id", "label")]
    for c in feat_cols:
        median = df[c].median()
        if pd.notna(median):
            df[c] = df[c].fillna(median)
        else:
            df[c] = df[c].fillna(0.0)

    feat_path = OUT_DIR / "pyradiomics_features.csv"
    df.to_csv(feat_path, index=False)
    print(f"  Feature matrix: {df.shape[0]} × {df.shape[1]-2} features → {feat_path}")

    return df


# ═══════════════════════════════════════════════════════════════
# 5. Machine Learning (Ruan 2025 method)
# ═══════════════════════════════════════════════════════════════

def run_ruan_ml(feature_df: pd.DataFrame):
    """LASSO + LRC/RFC/SVC with 10x5 CV."""
    from scipy import stats
    from sklearn.linear_model import LassoCV, LogisticRegression
    from sklearn.ensemble import RandomForestClassifier
    from sklearn.svm import SVC
    from sklearn.preprocessing import StandardScaler
    from sklearn.model_selection import StratifiedKFold
    from sklearn.metrics import (roc_auc_score, accuracy_score, f1_score,
                                  precision_score, recall_score, brier_score_loss)

    print("\n" + "=" * 70)
    print("RUAN 2025 ML PIPELINE")
    print("=" * 70)

    feat_cols = [c for c in feature_df.columns if c not in ("patient_id", "label")]
    X_all = feature_df[feat_cols].values.astype(np.float64)
    y = feature_df["label"].values.astype(int)

    print(f"\nData: {len(y)} patients, {len(feat_cols)} features")
    print(f"  PH={int((y==1).sum())}, nonPH={int((y==0).sum())}")

    # Step 1: Mann-Whitney U pre-filter
    print("\n[Step 1] Mann-Whitney U pre-filter (p<0.05)...")
    keep_idx = []
    pvals = []
    for j in range(X_all.shape[1]):
        g0 = X_all[y == 0, j]; g1 = X_all[y == 1, j]
        if np.std(np.r_[g0, g1]) < 1e-12:
            continue
        _, p = stats.mannwhitneyu(g0, g1, alternative="two-sided")
        if p < 0.05:
            keep_idx.append(j)
            pvals.append(p)

    X_mwu = X_all[:, keep_idx]
    mwu_names = [feat_cols[i] for i in keep_idx]
    print(f"  {len(keep_idx)}/{len(feat_cols)} features retained")

    if len(keep_idx) < 5:
        print("  WARNING: too few features after MWU, using top 50 by variance")
        var_idx = np.argsort(np.var(X_all, axis=0))[-50:]
        X_mwu = X_all[:, var_idx]
        mwu_names = [feat_cols[i] for i in var_idx]

    # Step 2: LASSO selection
    print("\n[Step 2] LASSO feature selection...")
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_mwu)
    alphas = np.logspace(-3, 1, 100, base=10)
    lasso = LassoCV(alphas=alphas, cv=10, max_iter=2000, random_state=SEED)
    lasso.fit(X_scaled, y)

    sel_mask = np.abs(lasso.coef_) > np.percentile(np.abs(lasso.coef_), 85)
    n_sel = int(sel_mask.sum())
    selected = [mwu_names[i] for i in range(len(mwu_names)) if sel_mask[i]]
    print(f"  LASSO selected {n_sel} features")

    # Build feature sets
    X_rad = X_mwu[:, sel_mask]  # Radiomic Model

    # Extract clinical features (age, WBC, etc) if available
    clin_cols = [c for c in feat_cols if c.startswith("clin_")]
    X_clin = feature_df[clin_cols].values.astype(np.float64) if clin_cols else np.zeros((len(y), 1))
    X_joint = np.hstack([X_rad, X_clin])  # Joint Model

    print(f"  RM features: {X_rad.shape[1]}, JM features: {X_joint.shape[1]}")

    # Step 3: Models
    print("\n[Step 3] Training models (10×5 CV)...")

    models = {
        "LRC_RM": LogisticRegression(penalty='l2', C=1.0, solver='lbfgs',
                                      max_iter=5000, class_weight='balanced',
                                      random_state=SEED),
        "RFC_RM": RandomForestClassifier(n_estimators=12, max_depth=4,
                                          max_features=11, random_state=SEED),
        "SVC_RM": SVC(kernel='rbf', C=1.0, gamma='scale', probability=True,
                       class_weight='balanced', random_state=SEED),
    }

    results = {}
    for mk, model in models.items():
        is_jm = mk.endswith("_JM")
        for suffix, X_use in [("_RM", X_rad), ("_JM", X_joint)]:
            name = mk.replace("_RM", suffix)
            if mk.endswith("_RM") and suffix == "_JM":
                model_jm = model.__class__(**model.get_params())
                name = mk.replace("_RM", "_JM")

            aucs = []
            for rep in range(N_REPEATS):
                skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True,
                                      random_state=SEED + rep * 100)
                for tr, va in skf.split(X_use, y):
                    X_tr = StandardScaler().fit_transform(X_use[tr])
                    X_va = StandardScaler().fit_transform(X_use[va])
                    X_tr_2 = StandardScaler().fit_transform(X_tr)
                    X_va_2 = StandardScaler().fit_transform(X_va)

                    if suffix == "_JM":
                        model_jm.fit(X_tr_2, y[tr])
                        try: yp = model_jm.predict_proba(X_va_2)[:, 1]
                        except: yp = model_jm.decision_function(X_va_2)
                    else:
                        model.fit(X_tr_2, y[tr])
                        try: yp = model.predict_proba(X_va_2)[:, 1]
                        except: yp = model.decision_function(X_va_2)

                    if len(set(y[va])) > 1:
                        aucs.append(roc_auc_score(y[va], yp))

            auc_arr = np.array(aucs)
            ci_lo, ci_hi = np.percentile(auc_arr, [2.5, 97.5])
            results[name] = {
                "AUC_mean": float(np.mean(auc_arr)),
                "AUC_std": float(np.std(auc_arr)),
                "AUC_CI_low": float(ci_lo),
                "AUC_CI_high": float(ci_hi),
                "n_folds": len(auc_arr),
            }
            print(f"  {name:12s}: AUC={results[name]['AUC_mean']:.4f} "
                  f"[{results[name]['AUC_CI_low']:.4f}-{results[name]['AUC_CI_high']:.4f}]")

    # Save results
    output = {
        "method": "Ruan 2025 — PyRadiomics + LASSO + 10x5 CV",
        "features": {"total_raw": len(feat_cols), "mwu_filtered": len(keep_idx),
                     "lasso_selected": n_sel},
        "selected_features": selected,
        "results": {k: {kk: float(vv) if isinstance(vv, (np.floating, np.integer)) else vv
                       for kk, vv in v.items()} for k, v in results.items()},
    }
    with open(OUT_DIR / "ruan2025_pyradiomics_results.json", "w") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)

    # Summary
    print("\n" + "=" * 70)
    print("FINAL RESULTS — Ruan 2025 PyRadiomics Method")
    print("=" * 70)
    for name, r in sorted(results.items(), key=lambda x: -x[1]["AUC_mean"]):
        print(f"  {name:12s}: AUC={r['AUC_mean']:.4f} [{r['AUC_CI_low']:.4f}-{r['AUC_CI_high']:.4f}]")

    return results


# ═══════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════

def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--skip-convert", action="store_true", help="Skip DICOM conversion")
    parser.add_argument("--skip-extract", action="store_true", help="Skip radiomics extraction")
    parser.add_argument("--skip-ml", action="store_true", help="Skip ML training")
    parser.add_argument("--max-patients", type=int, default=0, help="Limit patients (0=all)")
    args = parser.parse_args()

    t_start = time.time()
    print("=" * 70)
    print("COPD-PH PyRadiomics Pipeline — Ruan 2025 Method")
    print("=" * 70)

    # 1. Labels
    print("\n[PHASE 1] Building labels...")
    labels = build_labels()
    matched = find_ct_dirs(labels)
    if args.max_patients > 0:
        matched = matched[:args.max_patients]

    if not matched:
        print("ERROR: No patients matched with CT data!")
        print("Check CT data paths in find_ct_dirs()")
        return

    # 2. DICOM → NIfTI
    if not args.skip_convert:
        print(f"\n[PHASE 2] DICOM → NIfTI ({len(matched)} patients)...")
        t1 = time.time()
        convert_all(matched)
        print(f"  Time: {(time.time()-t1)/60:.1f} min")
    else:
        print("\n[PHASE 2] DICOM conversion SKIPPED")

    # 3. PyRadiomics extraction
    if not args.skip_extract:
        print(f"\n[PHASE 3] PyRadiomics feature extraction...")
        t2 = time.time()
        df = extract_all_radiomics(matched)
        print(f"  Time: {(time.time()-t2)/60:.1f} min")
    else:
        feat_path = OUT_DIR / "pyradiomics_features.csv"
        if feat_path.exists():
            df = pd.read_csv(feat_path)
            print(f"\n[PHASE 3] Loaded existing features: {df.shape}")
        else:
            print(f"\n[PHASE 3] ERROR: {feat_path} not found!")
            return

    # 4. ML training
    if not args.skip_ml:
        print(f"\n[PHASE 4] ML training (Ruan 2025 method)...")
        t3 = time.time()
        run_ruan_ml(df)
        print(f"  Time: {(time.time()-t3)/60:.1f} min")
    else:
        print("\n[PHASE 4] ML training SKIPPED")

    total = (time.time() - t_start) / 60
    print(f"\n{'=' * 70}")
    print(f"PIPELINE COMPLETE — {total:.1f} min total")
    print(f"Results: {OUT_DIR}")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
