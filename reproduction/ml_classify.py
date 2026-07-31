"""
COPD-PH vs COPD-nonPH ML Classification
=======================================
Reproduces the machine-learning classification from the GCN-copd-ph-classify repo.
Extracts quantitative CT + clinical features from Excel, trains multiple classifiers,
reports 6 standard metrics with stratified 5-fold CV.

Models: LogisticRegression, RandomForest, XGBoost, SVM, MLP (radiomics_only mode)
Feature sets: CT-only, Clinical-only, Combined
"""
from __future__ import annotations

import json
import os
import sys
import time
import warnings
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import openpyxl
import pandas as pd
from matplotlib import pyplot as plt
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score, auc, confusion_matrix, f1_score,
    precision_score, recall_score, roc_auc_score, roc_curve,
)
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.neural_network import MLPClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.svm import SVC

warnings.filterwarnings("ignore")

# ═══════════════════════════════════════════════════════════════════════
# Paths
# ═══════════════════════════════════════════════════════════════════════
ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "outputs" / "ml_classification"
OUT_DIR.mkdir(parents=True, exist_ok=True)

EXCEL = Path(r"E:\桌面文件\资料集-基于大模型与多智能体的COPD-PH资料集\copd-ph患者113例0331.xlsx")

SEED = 42
N_FOLDS = 5


# ═══════════════════════════════════════════════════════════════════════
# 1. Parse Excel → Feature Matrix
# ═══════════════════════════════════════════════════════════════════════

def parse_excel_data(excel_path: Path) -> pd.DataFrame:
    """Parse Excel into patient × features DataFrame with labels."""
    wb = openpyxl.load_workbook(str(excel_path))
    ws = wb.active

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # Build column name → 1-indexed column number map
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h)] = i + 1

    # Find key columns
    ph_col = col_map.get("PH", 7)
    name_col = col_map.get("name", 3)
    mpap_col = col_map.get("mPAP", 14)
    age_col = col_map.get("Age", 6)
    seg_col = next((i+1 for i, h in enumerate(headers) if h and "分割" in str(h)), 12)

    # ── Define feature groups ──
    # CT features (columns 250-424): quantitative CT radiomics
    ct_feature_ranges = [(250, 424)]  # all CT quantitative features

    # ⚠️  mPAP/PVSP/PA/PVR are DIAGNOSTIC gold-standard — including them = label leakage
    # ✅  Safe (non-leaking) features: demographics + labs + PFT + echo structure
    clinical_cols = {
        # — Demographics
        "age": 6, "BMI": 105, "GOLD": 99, "smoking": 102,
        # — Lab (non-PH-specific)
        "WBC": 42, "RBC": 43, "HB": 44, "PLT": 45,
        "ALT": 46, "AST": 47, "Crea": 51, "UA": 52, "D-Dimer": 60,
        "cTnT": 28,
        # — PFT
        "FVC%": 171, "FEV1%": 174, "FEV1FVC": 175,
        "DLCO_SB%": 192, "DLCO_VA%": 195,
        "TLC%": 201, "RV%": 199, "RVTLC": 206,
        # — Echo STRUCTURE (not pressure)
        "LVEF": 82, "TAPSE": 81, "RVTD": 79, "RVLD": 80, "RVWDd": 87,
        "Aortic": 94,
    }
    # ❌ LEAKAGE (explicitly excluded):
    #   mPAP(14), PVSP(83), PA(22), PA_diameter(88), PVR(19),
    #   mRAP(15), mPAWP(16), CO(17), CI(18), NT-ProBNP(27), 6MWD(40)

    # ── Build patient records ──
    records = []
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col).value
        ph = ws.cell(row=row, column=ph_col).value
        age = ws.cell(row=row, column=age_col).value
        seg = ws.cell(row=row, column=seg_col).value

        # Label
        if ph == "是":
            label = 1
        elif ph == "/":
            label = 0
        else:
            continue  # unknown

        rec = {
            "name": name,
            "label": label,
            "seg_folder": str(seg).rstrip("/").rstrip("\\") if seg else "",
        }

        # Clinical features
        for fname, ci in clinical_cols.items():
            val = ws.cell(row=row, column=ci).value
            try:
                rec[f"clin_{fname}"] = float(val)
            except (ValueError, TypeError):
                rec[f"clin_{fname}"] = np.nan

        # CT quantitative features
        for ci in range(250, ws.max_column + 1):
            h = headers[ci - 1]
            if h is None:
                continue
            val = ws.cell(row=row, column=ci).value
            try:
                rec[f"ct_{ci}"] = float(val)
            except (ValueError, TypeError):
                rec[f"ct_{ci}"] = np.nan

        records.append(rec)

    df = pd.DataFrame(records)
    print(f"Parsed {len(df)} patients: PH={df['label'].sum()}, nonPH={(df['label']==0).sum()}")
    return df


def prepare_feature_sets(df: pd.DataFrame) -> Dict[str, Tuple[np.ndarray, np.ndarray, List[str]]]:
    """Build X (features) and y (labels) for different feature sets.

    Returns dict: feature_set_name → (X, y, feature_names)
    """
    y = df["label"].values.astype(int)

    # ── CT-only features ──
    ct_cols = [c for c in df.columns if c.startswith("ct_")]
    X_ct = df[ct_cols].values.astype(np.float32)

    # Keep columns with >50% non-NaN
    ct_valid = ~np.all(np.isnan(X_ct), axis=0)
    ct_nan_rate = np.isnan(X_ct).mean(axis=0)
    ct_keep = ct_nan_rate < 0.5
    X_ct = X_ct[:, ct_keep]
    ct_names = [ct_cols[i] for i in range(len(ct_cols)) if ct_keep[i]]

    # Impute NaN with median
    for j in range(X_ct.shape[1]):
        col = X_ct[:, j]
        mask = np.isnan(col)
        if mask.any():
            col[mask] = np.nanmedian(col)

    print(f"  CT features: {X_ct.shape[1]} (kept from {len(ct_cols)}, "
          f"dropped {sum(~ct_keep)} with >50% NaN)")

    # ── Clinical-only features ──
    clin_cols = [c for c in df.columns if c.startswith("clin_")]
    X_clin = df[clin_cols].values.astype(np.float32)

    clin_valid = []
    clin_names = []
    for j, cn in enumerate(clin_cols):
        col = X_clin[:, j]
        nan_rate = np.isnan(col).mean()
        if nan_rate < 0.7:  # keep if <70% NaN
            clin_valid.append(j)
            clin_names.append(cn)

    X_clin = X_clin[:, clin_valid]

    # Impute
    for j in range(X_clin.shape[1]):
        col = X_clin[:, j]
        mask = np.isnan(col)
        if mask.any():
            col[mask] = np.nanmedian(col)

    print(f"  Clinical features: {X_clin.shape[1]} (from {len(clin_cols)})")

    # ── Combined ──
    X_comb = np.hstack([X_ct, X_clin])
    comb_names = ct_names + clin_names

    # ── Top-K CT features (for interpretable model) ──
    # Use simple variance-based selection: top 30 highest-variance CT features
    ct_var = np.nanvar(X_ct, axis=0)
    top_k = min(30, X_ct.shape[1])
    top_idx = np.argsort(ct_var)[-top_k:]
    X_ct_top = X_ct[:, top_idx]
    ct_top_names = [ct_names[i] for i in top_idx]

    return {
        "CT_all": (X_ct, y, ct_names),
        "CT_top30": (X_ct_top, y, ct_top_names),
        "Clinical_safe": (X_clin, y, clin_names),
        "Combined_safe": (X_comb, y, comb_names),
    }


# ═══════════════════════════════════════════════════════════════════════
# 2. Training & Evaluation
# ═══════════════════════════════════════════════════════════════════════

def specificity_score(y_true, y_pred):
    tn = ((y_true == 0) & (y_pred == 0)).sum()
    fp = ((y_true == 0) & (y_pred == 1)).sum()
    return float(tn / (tn + fp)) if (tn + fp) > 0 else 0.0


def all_metrics(y_true, y_pred, y_prob=None):
    m = {
        "AUC": roc_auc_score(y_true, y_prob) if y_prob is not None and len(set(y_true)) > 1 else 0.0,
        "Accuracy": accuracy_score(y_true, y_pred),
        "Precision": precision_score(y_true, y_pred, zero_division=0),
        "Sensitivity": recall_score(y_true, y_pred, zero_division=0),
        "F1": f1_score(y_true, y_pred, zero_division=0),
        "Specificity": specificity_score(y_true, y_pred),
    }
    return m


def run_cv_experiment(
    X: np.ndarray,
    y: np.ndarray,
    model_name: str,
    model,
    feature_set_name: str,
    need_scaling: bool = False,
) -> dict:
    """Run stratified 5-fold CV for one model."""
    skf = StratifiedKFold(n_splits=N_FOLDS, shuffle=True, random_state=SEED)

    fold_results = []
    all_yt, all_yp, all_pr = [], [], []

    for fold_i, (train_idx, val_idx) in enumerate(skf.split(X, y)):
        X_tr, X_val = X[train_idx], X[val_idx]
        y_tr, y_val = y[train_idx], y[val_idx]

        if need_scaling:
            scaler = StandardScaler()
            X_tr = scaler.fit_transform(X_tr)
            X_val = scaler.transform(X_val)
        else:
            scaler = None

        model_clone = model.__class__(**model.get_params())
        model_clone.fit(X_tr, y_tr)

        y_pred = model_clone.predict(X_val)
        try:
            y_prob = model_clone.predict_proba(X_val)[:, 1]
        except (AttributeError, NotImplementedError):
            y_prob = model_clone.decision_function(X_val)
            y_prob = (y_prob - y_prob.min()) / (y_prob.max() - y_prob.min() + 1e-8)

        m = all_metrics(y_val, y_pred, y_prob)
        m["fold"] = fold_i
        fold_results.append(m)

        all_yt.extend(y_val.tolist())
        all_yp.extend(y_pred.tolist())
        all_pr.extend(y_prob.tolist())

    # Aggregate
    all_yt = np.array(all_yt)
    all_yp = np.array(all_yp)
    all_pr = np.array(all_pr)

    keys = ["AUC", "Accuracy", "Precision", "Sensitivity", "F1", "Specificity"]
    mean_vals = {k: float(np.mean([f[k] for f in fold_results])) for k in keys}
    std_vals  = {k: float(np.std([f[k] for f in fold_results])) for k in keys}
    pooled_auc = float(roc_auc_score(all_yt, all_pr)) if len(set(all_yt)) > 1 else 0.0

    return {
        "model": model_name,
        "feature_set": feature_set_name,
        "folds": fold_results,
        "mean": mean_vals,
        "std": std_vals,
        "pooled_AUC": pooled_auc,
        "n_samples": len(y),
        "n_ph": int((y == 1).sum()),
        "n_nonph": int((y == 0).sum()),
        # For ROC plotting
        "y_true": all_yt.tolist(),
        "y_prob": all_pr.tolist(),
    }


# ═══════════════════════════════════════════════════════════════════════
# 3. Visualizations
# ═══════════════════════════════════════════════════════════════════════

def make_plots(all_results: List[dict]):
    """Generate ROC curves and comparison plots."""
    # ── ROC curves ──
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    # By model (Combined features)
    ax = axes[0]
    colors = plt.cm.Set2(np.linspace(0, 1, 8))
    for i, r in enumerate(all_results):
        if r["feature_set"] != "Combined":
            continue
        fpr, tpr, _ = roc_curve(r["y_true"], r["y_prob"])
        roc_auc = auc(fpr, tpr)
        ax.plot(fpr, tpr, color=colors[i % len(colors)], lw=2,
                label=f"{r['model']} (AUC={roc_auc:.3f})")
    ax.plot([0, 1], [0, 1], 'k--', alpha=0.3)
    ax.set_xlabel("1 - Specificity")
    ax.set_ylabel("Sensitivity")
    ax.set_title("ROC Curves — Combined Features (mPAP/PVSP excluded)")
    ax.legend(loc="lower right", fontsize=8)

    # By feature set (best model)
    ax = axes[1]
    best_model = max(all_results, key=lambda r: r["mean"]["AUC"])
    best_name = best_model["model"]
    feature_sets = ["Clinical_safe", "CT_top30", "CT_all", "Combined_safe"]
    colors2 = plt.cm.viridis(np.linspace(0.2, 0.9, len(feature_sets)))
    for j, fs in enumerate(feature_sets):
        matches = [r for r in all_results if r["model"] == best_name and r["feature_set"] == fs]
        if not matches:
            matches = [r for r in all_results if r["feature_set"] == fs]  # fallback
            if not matches:
                continue
        r = matches[0]
        fpr, tpr, _ = roc_curve(r["y_true"], r["y_prob"])
        roc_auc = auc(fpr, tpr)
        ax.plot(fpr, tpr, color=colors2[j], lw=2,
                label=f"{fs} (AUC={roc_auc:.3f})")
    ax.plot([0, 1], [0, 1], 'k--', alpha=0.3)
    ax.set_xlabel("1 - Specificity")
    ax.set_ylabel("Sensitivity")
    ax.set_title(f"ROC by Feature Set ({r.get('model', best_name)})")
    ax.legend(loc="lower right", fontsize=8)

    plt.tight_layout()
    roc_path = OUT_DIR / "roc_curves.png"
    fig.savefig(roc_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"ROC curves → {roc_path}")

    # ── Metrics bar chart ──
    fig, ax = plt.subplots(figsize=(14, 7))
    metrics_keys = ["AUC", "Accuracy", "Sensitivity", "Specificity", "F1", "Precision"]
    x = np.arange(len(metrics_keys))
    width = 0.12

    # Group by model (Combined_safe features)
    comb_results = [r for r in all_results if r["feature_set"] == "Combined_safe"]
    for i, r in enumerate(comb_results):
        vals = [r["mean"][k] for k in metrics_keys]
        errs = [r["std"][k] for k in metrics_keys]
        bars = ax.bar(x + i * width, vals, width, label=r["model"],
                      yerr=errs, capsize=3, color=colors[i % len(colors)])
        # Annotate
        for bar, val in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 0.01,
                    f'{val:.3f}', ha='center', va='bottom', fontsize=6, rotation=90)

    ax.set_ylabel("Score")
    ax.set_title("Model Comparison — Combined Features (mPAP/PVSP/PA excluded)")
    ax.set_xticks(x + width * (len(comb_results) - 1) / 2)
    ax.set_xticklabels(metrics_keys)
    ax.set_ylim(0, 1.15)
    ax.legend(fontsize=8, loc="lower right")
    ax.grid(axis='y', alpha=0.3)
    plt.tight_layout()
    bar_path = OUT_DIR / "metrics_comparison.png"
    fig.savefig(bar_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    print(f"Metrics bar chart → {bar_path}")


# ═══════════════════════════════════════════════════════════════════════
# 4. Main
# ═══════════════════════════════════════════════════════════════════════

def main():
    t0 = time.time()
    print("=" * 70)
    print("COPD-PH ML Classification — Reproduction")
    print("=" * 70)

    # ── Parse data ──
    print("\n[1/4] Parsing Excel data...")
    df = parse_excel_data(EXCEL)

    # Remove patients without any CT data
    ct_cols = [c for c in df.columns if c.startswith("ct_")]
    has_ct = df[ct_cols].notna().any(axis=1)
    df = df[has_ct].copy()
    print(f"  After removing patients without CT data: {len(df)} patients "
          f"(PH={df['label'].sum()}, nonPH={(df['label']==0).sum()})")

    if len(df) < 20:
        print("ERROR: Too few patients with CT data!")
        return 1

    # ── Prepare features ──
    print("\n[2/4] Building feature sets...")
    feature_sets = prepare_feature_sets(df)

    # ── Define models ──
    models = {
        "LogisticRegression": LogisticRegression(
            penalty='l2', C=1.0, solver='lbfgs', max_iter=5000,
            class_weight='balanced', random_state=SEED,
        ),
        "RandomForest": RandomForestClassifier(
            n_estimators=200, max_depth=8, min_samples_leaf=5,
            class_weight='balanced', random_state=SEED, n_jobs=-1,
        ),
        "XGBoost": None,  # conditional
        "SVM_RBF": SVC(
            kernel='rbf', C=1.0, gamma='scale', class_weight='balanced',
            probability=True, random_state=SEED,
        ),
        "MLP": MLPClassifier(
            hidden_layer_sizes=(64, 32), activation='relu', alpha=0.001,
            batch_size=8, learning_rate='adaptive', learning_rate_init=1e-3,
            max_iter=500, early_stopping=True, random_state=SEED,
        ),
    }

    # Try XGBoost
    try:
        import xgboost as xgb
        scale_pos_weight = (df['label'] == 0).sum() / max((df['label'] == 1).sum(), 1)
        models["XGBoost"] = xgb.XGBClassifier(
            n_estimators=200, max_depth=5, learning_rate=0.05,
            scale_pos_weight=scale_pos_weight,
            subsample=0.8, colsample_bytree=0.8,
            random_state=SEED, eval_metric='logloss',
        )
    except ImportError:
        print("  XGBoost not available, skipping")

    need_scaling_models = {"LogisticRegression", "SVM_RBF", "MLP"}

    # ── Run all experiments ──
    print("\n[3/4] Running cross-validation experiments...")
    all_results = []

    for fs_name, (X, y, feat_names) in feature_sets.items():
        print(f"\n{'─' * 50}")
        print(f"Feature set: {fs_name} ({X.shape[1]} features, n={len(y)})")
        print(f"{'─' * 50}")

        for model_name, model in models.items():
            if model is None:
                continue

            need_scaling = model_name in need_scaling_models
            t1 = time.time()
            result = run_cv_experiment(
                X, y, model_name, model, fs_name,
                need_scaling=need_scaling,
            )
            elapsed = time.time() - t1

            m = result["mean"]
            print(f"  {model_name:<25s} "
                  f"AUC={m['AUC']:.4f} Acc={m['Accuracy']:.4f} "
                  f"Sens={m['Sensitivity']:.4f} Spec={m['Specificity']:.4f} "
                  f"F1={m['F1']:.4f} — {elapsed:.1f}s")

            all_results.append(result)

    # ── Save & Report ──
    print(f"\n[4/4] Saving results and generating visualizations...")

    # Save JSON results
    results_json = {
        "config": {
            "n_folds": N_FOLDS, "seed": SEED,
            "total_patients": int(len(df)),
            "n_ph": int(df['label'].sum()),
            "n_nonph": int((df['label'] == 0).sum()),
        },
        "results": [],
    }
    for r in all_results:
        entry = {k: v for k, v in r.items() if k not in ("y_true", "y_prob")}
        results_json["results"].append(entry)

    json_path = OUT_DIR / "ml_results.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results_json, f, indent=2, ensure_ascii=False)
    print(f"Results JSON → {json_path}")

    # Best result
    best = max(all_results, key=lambda r: r["mean"]["AUC"])
    print(f"\n  ★ BEST: {best['model']} + {best['feature_set']}")
    print(f"    AUC={best['mean']['AUC']:.4f} ± {best['std']['AUC']:.4f}")

    # ── Summary table ──
    print(f"\n{'=' * 110}")
    print("FINAL COMPARISON TABLE")
    print(f"{'=' * 110}")
    header = (f"{'Model':<25s} {'Features':<14s} {'AUC':>8s} {'Acc':>8s} "
              f"{'Sens':>8s} {'Spec':>8s} {'F1':>8s} {'Prec':>8s}  pooled_AUC")
    print(header)
    print("-" * 110)
    for r in sorted(all_results, key=lambda r: -r["mean"]["AUC"]):
        m = r["mean"]
        print(f"{r['model']:<25s} {r['feature_set']:<14s} "
              f"{m['AUC']:8.4f} {m['Accuracy']:8.4f} "
              f"{m['Sensitivity']:8.4f} {m['Specificity']:8.4f} "
              f"{m['F1']:8.4f} {m['Precision']:8.4f}  {r['pooled_AUC']:.4f}")
    print(f"{'=' * 110}")

    # ── Plots ──
    make_plots(all_results)

    # ── Excel report ──
    summary_rows = []
    for r in all_results:
        m = r["mean"]
        s = r["std"]
        summary_rows.append({
            "Model": r["model"],
            "Features": r["feature_set"],
            "AUC": f"{m['AUC']:.4f}±{s['AUC']:.4f}",
            "Accuracy": f"{m['Accuracy']:.4f}±{s['Accuracy']:.4f}",
            "Sensitivity": f"{m['Sensitivity']:.4f}±{s['Sensitivity']:.4f}",
            "Specificity": f"{m['Specificity']:.4f}±{s['Specificity']:.4f}",
            "F1": f"{m['F1']:.4f}±{s['F1']:.4f}",
            "Precision": f"{m['Precision']:.4f}±{s['Precision']:.4f}",
            "pooled_AUC": f"{r['pooled_AUC']:.4f}",
            "n_ph": r["n_ph"],
            "n_nonph": r["n_nonph"],
        })
    summary_df = pd.DataFrame(summary_rows)
    xlsx_path = OUT_DIR / "ml_results_summary.xlsx"
    summary_df.to_excel(xlsx_path, index=False)
    print(f"Excel report → {xlsx_path}")

    elapsed = time.time() - t0
    print(f"\nTotal time: {elapsed/60:.1f} min")
    return 0


if __name__ == "__main__":
    sys.exit(main())
